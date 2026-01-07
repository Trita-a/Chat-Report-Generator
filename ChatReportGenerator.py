import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import sys
import os
import re
import argparse
import shutil
from datetime import datetime
import openpyxl

# ==========================================
# CSS STYLES
# ==========================================

CSS_SIGNAL = """
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Oxygen, Ubuntu, Cantarell, "Open Sans", "Helvetica Neue", sans-serif; background-color: #f7f7f7; margin: 0; padding: 0; color: #1b1b1b; }
.container { display: flex; height: 100vh; }
.sidebar { width: 320px; flex-shrink: 0; background-color: white; border-right: 1px solid #e6e6e6; overflow-y: auto; }
.main { flex: 1; display: flex; flex-direction: column; background-color: #ffffff; }
.chat-messages { flex: 1; overflow-y: auto; padding: 20px 40px; display: flex; flex-direction: column; }
.message { max-width: 60%; margin-bottom: 8px; padding: 10px 14px; border-radius: 16px; position: relative; font-size: 15px; line-height: 1.4; word-wrap: break-word; }
.message.received { align-self: flex-start; background-color: #f2f2f2; color: #1b1b1b; border-bottom-left-radius: 4px; margin-right: auto; }
.message.sent { align-self: flex-end; background-color: #2c6bed; color: white; border-bottom-right-radius: 4px; margin-left: auto; }
.sender-name { font-size: 12px; font-weight: bold; margin-bottom: 2px; color: #666; }
.message.sent .sender-name { display: none; }
.timestamp { font-size: 10px; opacity: 0.6; float: right; margin-left: 8px; margin-top: 4px; }
.message.sent .timestamp { color: rgba(255,255,255,0.8); }

/* SIDEBAR STYLES */
.signal-sidebar-header { font-size: 20px; font-weight: bold; padding: 24px 20px 16px 20px; border-bottom: 1px solid #e6e6e6; color: #1b1b1b; }
.chat-item { height: 72px; padding: 0 20px; border-bottom: 1px solid #f5f5f5; cursor: pointer; display: flex; flex-direction: row; align-items: center; transition: background-color 0.1s; }
.chat-item:hover { background-color: #f5f5f5; }
.chat-item.active { background-color: #d1e1fb; border-left: 4px solid #2c6bed; }
.sidebar-avatar { width: 52px; height: 52px; font-size: 20px; margin-right: 16px; margin-bottom: 0; display: flex; align-items: center; justify-content: center; background-color: #ddd; border-radius: 50%; color: #333; font-weight: bold; }
.sidebar-name { font-weight: 600; font-size: 15px; color: #1b1b1b; }
.sidebar-info { font-size: 13px; color: #666; margin-top: 3px; }

/* SHARED / OTHER */
.chat-info-header { background-color: #f7f7f7; padding: 15px; border-bottom: 1px solid #ddd; display: flex; justify-content: center; align-items: center; gap: 40px; margin: 10px 20px 20px 20px; border-radius: 8px; }
.participant-avatar { width: 60px; height: 60px; border-radius: 50%; background-color: #ddd; color: #333; display: flex; align-items: center; justify-content: center; font-size: 24px; font-weight: bold; margin-bottom: 8px; }
.participant-card { display: flex; flex-direction: column; align-items: center; text-align: center; }
.participant-name { font-weight: 700; font-size: 16px; color: #1b1b1b; }
.participant-number { font-size: 13px; color: #666; margin-top: 2px; }
.translation { margin-top: 8px; padding: 8px 10px; border-left: 3px solid #fecb00; font-style: normal; font-size: 13.5px; background-color: rgba(255,255,255,0.95); color: #333; border-radius: 6px; display: block; clear: both; margin-bottom: 2px; }
.date-divider { align-self: center; color: #555; font-size: 13px; margin: 15px 0; font-weight: 500; text-align: center; }
.message.sent .translation { background-color: rgba(255,255,255,0.15); color: white; border-left-color: rgba(255,255,255,0.5); }
.attachment img { max-width: 250px; border-radius: 12px; margin-top: 5px; cursor: pointer; }
.attachment video { max-width: 250px; border-radius: 12px; margin-top: 5px; }
.attachment audio { max-width: 250px; margin-top: 5px; }
.search-box { padding: 10px 15px; background-color: #f7f7f7; border-bottom: 1px solid #e6e6e6; }
.search-input { width: 100%; padding: 8px 12px; border-radius: 6px; border: 1px solid #ddd; background-color: #e9e9e9; font-size: 14px; box-sizing: border-box; outline: none; }
.search-input:focus { background-color: white; border-color: #2c6bed; }

/* ADVANCED SEARCH - PREMIUM */
.search-results-container { display: none; flex: 1; overflow-y: auto; background-color: #fff; }
.search-header { padding: 16px 20px; font-weight: 700; font-size: 14px; color: #1b1b1b; background: #f7f7f7; border-bottom: 1px solid #e6e6e6; display: flex; justify-content: space-between; align-items: center; position: sticky; top: 0; z-index: 5; }
.close-search { cursor: pointer; color: #2c6bed; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; padding: 4px 8px; border-radius: 4px; transition: background 0.2s; }
.close-search:hover { background-color: rgba(44, 107, 237, 0.1); }
.search-result-item { padding: 12px 20px; border-bottom: 1px solid #f5f5f5; cursor: pointer; transition: background 0.2s; display: flex; align-items: center; }
.search-result-item:hover { background-color: #f0f2f5; }
.search-avatar { width: 42px; height: 42px; border-radius: 50%; background-color: #e0e0e0; color: #555; display: flex; align-items: center; justify-content: center; font-weight: bold; font-size: 16px; margin-right: 14px; flex-shrink: 0; }
.search-content { flex: 1; min-width: 0; }
.search-result-sender { font-size: 14px; font-weight: 600; color: #1b1b1b; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.search-result-preview { font-size: 13px; color: #666; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; line-height: 1.4; }
.highlight-term { background-color: rgba(255, 235, 59, 0.5); border-radius: 2px; color: #000; font-weight: 500; padding: 0 2px; }
.highlight-msg { position: relative; animation: flashMsg 2s ease-out; }
.highlight-msg::after { content: ""; position: absolute; top: 0; left: 0; right: 0; bottom: 0; background-color: rgba(255, 235, 59, 0.2); border-radius: inherit; pointer-events: none; opacity: 0; animation: fadeOverlay 3s forwards; }
@keyframes flashMsg { 0% { transform: scale(1.02); box-shadow: 0 4px 12px rgba(0,0,0,0.1); z-index: 10; } 100% { transform: scale(1); box-shadow: none; z-index: 1; } }
@keyframes fadeOverlay { 0% { opacity: 1; } 100% { opacity: 0; } }
"""

CSS_WHATSAPP = """
body { font-family: "Segoe UI", Roboto, Helvetica, Arial, sans-serif; background-color: #d1d7db; margin: 0; padding: 0; color: #111b21; }
.container { display: flex; height: 100vh; }
.sidebar { width: 350px; flex-shrink: 0; background-color: white; border-right: 1px solid #e9edef; overflow-y: auto; }
.main { flex: 1; display: flex; flex-direction: column; background-color: #efe7dd; background-image: url("https://user-images.githubusercontent.com/15075759/28719144-86dc0f70-73b1-11e7-911d-60d70fcded21.png"); }
.chat-messages { flex: 1; overflow-y: auto; padding: 20px 60px; display: flex; flex-direction: column; }
.message { max-width: 65%; margin-bottom: 8px; padding: 6px 7px 8px 9px; border-radius: 7.5px; position: relative; font-size: 14.2px; line-height: 19px; box-shadow: 0 1px 0.5px rgba(11,20,26,0.13); word-wrap: break-word; }
.message.received { align-self: flex-start; background-color: #ffffff; color: #111b21; border-top-left-radius: 0; margin-right: auto; }
.message.sent { align-self: flex-end; background-color: #d9fdd3; color: #111b21; border-top-right-radius: 0; margin-left: auto; }
.timestamp { font-size: 11px; float: right; margin-left: 10px; margin-top: 5px; color: #667781; }
.chat-item { height: 72px; display: flex; align-items: center; padding: 0 15px; border-bottom: 1px solid #f0f2f5; cursor: pointer; flex-direction: row; }
.chat-item:hover { background-color: #f5f6f6; }
.chat-item.active { background-color: #ebebeb; border-left: 4px solid #00a884; }
.chat-info-header { background-color: rgba(255,255,255,0.92); padding: 15px; border-radius: 12px; margin: 20px auto; max-width: 600px; display: flex; justify-content: center; align-items: center; gap: 30px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }
.chat-header-bar { background-color: #f0f2f5; padding: 10px 16px; border-left: 1px solid #d1d7db; display: flex; align-items: center; height: 60px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); z-index: 10; }
.participant-avatar { width: 60px; height: 60px; border-radius: 50%; background-color: #e9edef; color: #fff; display: flex; align-items: center; justify-content: center; font-size: 24px; font-weight: bold; margin-bottom: 8px; border: 3px solid white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
.participant-avatar.owner { background-color: #00a884; }
.participant-avatar.contact { background-color: #555; }
.participant-card { display: flex; flex-direction: column; align-items: center; text-align: center; }
.participant-name { font-weight: 600; font-size: 15px; color: #111b21; }
.participant-number { font-size: 12px; color: #667781; }
.sidebar-avatar { width: 45px; height: 45px; font-size: 18px; margin-right: 15px; margin-bottom: 0; display: flex; align-items: center; justify-content: center; background-color: #e9edef; border-radius: 50%; color: #fff; font-weight: bold; }
.sidebar-name { font-weight: 600; font-size: 15px; color: #111b21; }
.sidebar-info { font-size: 13px; color: #667781; }
.translation { margin-top: 6px; padding: 6px 8px; border-left: 3px solid #ffcc00; font-size: 13px; background-color: rgba(0,0,0,0.03); color: #333; border-radius: 4px; font-style: italic; }
.attachment video { max-width: 100%; border-radius: 8px; margin-top: 5px; }
.attachment audio { max-width: 100%; margin-top: 5px; }
.search-box { padding: 10px; background-color: #f0f2f5; border-bottom: 1px solid #d1d7db; }
.search-input { width: 100%; padding: 7px 12px; border-radius: 8px; border: none; background-color: #ffffff; font-size: 14px; box-sizing: border-box; height: 35px; outline: none; }

/* ADVANCED SEARCH - PREMIUM */
.search-results-container { display: none; flex: 1; overflow-y: auto; background-color: #fff; }
.search-header { padding: 16px 20px; font-weight: 700; font-size: 14px; color: #1b1b1b; background: #f7f7f7; border-bottom: 1px solid #e6e6e6; display: flex; justify-content: space-between; align-items: center; position: sticky; top: 0; z-index: 5; }
.close-search { cursor: pointer; color: #2c6bed; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; padding: 4px 8px; border-radius: 4px; transition: background 0.2s; }
.close-search:hover { background-color: rgba(44, 107, 237, 0.1); }
.search-result-item { padding: 12px 20px; border-bottom: 1px solid #f5f5f5; cursor: pointer; transition: background 0.2s; display: flex; align-items: center; }
.search-result-item:hover { background-color: #f0f2f5; }
.search-avatar { width: 42px; height: 42px; border-radius: 50%; background-color: #e0e0e0; color: #555; display: flex; align-items: center; justify-content: center; font-weight: bold; font-size: 16px; margin-right: 14px; flex-shrink: 0; }
.search-content { flex: 1; min-width: 0; }
.search-result-sender { font-size: 14px; font-weight: 600; color: #1b1b1b; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.search-result-preview { font-size: 13px; color: #666; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; line-height: 1.4; }
.highlight-term { background-color: rgba(255, 235, 59, 0.5); border-radius: 2px; color: #000; font-weight: 500; padding: 0 2px; }
.highlight-msg { position: relative; animation: flashMsg 2s ease-out; }
.highlight-msg::after { content: ""; position: absolute; top: 0; left: 0; right: 0; bottom: 0; background-color: rgba(255, 235, 59, 0.2); border-radius: inherit; pointer-events: none; opacity: 0; animation: fadeOverlay 3s forwards; }
@keyframes flashMsg { 0% { transform: scale(1.02); box-shadow: 0 4px 12px rgba(0,0,0,0.1); z-index: 10; } 100% { transform: scale(1); box-shadow: none; z-index: 1; } }
@keyframes fadeOverlay { 0% { opacity: 1; } 100% { opacity: 0; } }
"""

# ==========================================
# HELPERS
# ==========================================

def clean_text(text):
    if text is None:
        return ""
    
    text = str(text)
    if not text or text.lower() == 'nan':
         return ""

    patterns = [
        r'Etichette:.*?(?=\n|$)',
        r'Creato:.*?(?=\n|$)',
        r'Modificato:.*?(?=\n|$)',
        r'Descrizione:.*?(?=\n|$)',
        r'Generator:.*?(?=\n|$)',
    ]
    cleaned = text
    for p in patterns:
        cleaned = re.sub(p, '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\n\s*\n', '\n', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\n\s*\n', '\n', cleaned).strip()
    cleaned = cleaned.replace("_x000d_", "")
    return cleaned

def parse_participants_intelligent(data, chat_id=""):
    """
    Intelligent logic to extract Owner and Contact from chat data.
    Returns: (owner_name, owner_number, contact_name, contact_number)
    """
    def extract_name_number(text_segment):
        cleaned = re.sub(r'\(proprietario\)', '', text_segment, flags=re.IGNORECASE)
        cleaned = re.sub(r'\(owner\)', '', cleaned, flags=re.IGNORECASE)
        cleaned = cleaned.replace("_x000d_", "").strip()
        
        number_match = re.search(r'\+?\d{8,}', cleaned)
        number = ""
        name = cleaned
        
        if number_match:
            number = number_match.group(0)
            name = cleaned.replace(number, "").strip()
        
        name = name.strip(",.- ")
        if not name: name = "Sconosciuto"
        if not number: number = "(Numero non presente)"
        
        return name, number

    owner_account_raw = str(data.get('owner', ''))
    owner_digits = re.sub(r'\D', '', owner_account_raw)
    
    parts_raw = str(data.get('participants', ''))
    segments = re.split(r'_x000d_|\n', parts_raw)
    
    owner_name = "Proprietario"
    owner_num = owner_account_raw
    
    contact_name = "Contatto"
    contact_num = ""
    
    found_contact = False
    
    # Fallback for space separated (WhatsApp)
    if ' ' in parts_raw and '_x000d_' not in parts_raw and '\n' not in parts_raw:
        segments = parts_raw.split(' ')

    for seg in segments:
        if not seg.strip(): continue
        
        s_name, s_number = extract_name_number(seg)
        s_digits = re.sub(r'\D', '', s_number)
        
        # Check if owner
        if owner_digits and len(owner_digits) > 5 and owner_digits in s_digits:
            owner_name = s_name
            owner_num = s_number
        else:
            if not found_contact:
                contact_name = s_name
                contact_num = s_number
                found_contact = True
            else:
                contact_name += f", {s_name}"

    if "Sconosciuto" in owner_name or not owner_name:
         owner_name = "Tu"

    if contact_name == "Contatto":
         if "Chat" not in str(chat_id):
             contact_name = str(chat_id)
             
    return owner_name, owner_num, contact_name, contact_num

def process_attachments(source_dir, output_dir):
    """
    Scans source_dir for files. 
    Copies them to output_dir/attachments.
    Returns mapping {filename: 'attachments/filename'}
    """
    mapping = {}
    att_dir = os.path.join(output_dir, "attachments")
    os.makedirs(att_dir, exist_ok=True)
    
    print(f"Scanning and copying attachments from {source_dir}...")
    
    exclude_prefix = os.path.abspath(output_dir)
    
    for root, dirs, files in os.walk(source_dir):
        if os.path.abspath(root).startswith(exclude_prefix):
            continue
            
        for file in files:
            # Skip the report itself or script
            if file in ["index.html", "generate_chat_report.py", "ChatReportGenerator.py", "chat_report_gui.py", "ChatReportGenerator_Pandas.py"]:
                continue
                
            src_path = os.path.join(root, file)
            # Flatten structure: copy all unique filenames to attachments/
            dest_path = os.path.join(att_dir, file)
            
            try:
                if not os.path.exists(dest_path):
                    shutil.copy2(src_path, dest_path)
                
                # Rel path for HTML
                mapping[file] = f"attachments/{file}"
            except Exception as e:
                pass
                
    return mapping

# ==========================================
# PARSERS (OpenPyXL Implementation)
# ==========================================

class BaseParser:
    def parse(self, filepath):
        raise NotImplementedError

class CellebriteParser(BaseParser):
    def parse(self, filepath):
        print("Using CellebriteParser (Light)...")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        sheet = None
        if 'Chat' in wb.sheetnames:
            sheet = wb['Chat']
        else:
            # Fallback to first
            sheet = wb.active
            
        chats = {}
        
        # Determine start row
        # Usually Cellebrite export has title then header. So row 1 is Title, Row 2 is Header, Row 3 is Data.
        # But indices in pandas were header=1 (so row 0 skip, row 1 header).
        # We will assume data starts from row 3 (index 3 in openpyxl) if checking pandas logic.
        # Let's start iterating from row 2 and see if it looks like data.
        
        # We'll skip first 2 rows.
        rows = list(sheet.iter_rows(values_only=True))
        
        # Heuristic: skip rows until we find one with an ID or looks like data
        # Pandas `header=1` means:
        # 0: "Report generated..."
        # 1: "ID", "Source", "Type"... (Header)
        # 2: 1, "WhatsApp", ... (Data)
        
        data_rows = rows[2:] # Slice off 0 and 1
        
        for idx, row in enumerate(data_rows):
            try:
                if not row or len(row) < 51: # Ensure row has minimal length
                    continue
                    
                # Index mapping (Pandas 0-based -> row[i])
                # 1: Chat ID
                chat_id = row[1]
                if chat_id is None: continue
                
                if chat_id not in chats:
                    chats[chat_id] = {
                        "id": chat_id,
                        "participants": str(row[9]) if row[9] is not None else "",
                        "owner": str(row[12]) if row[12] is not None else "",
                        "messages": []
                    }
                
                sender = str(row[21]) if row[21] is not None else ""
                owner_val = str(row[12]) if row[12] is not None else ""
                
                # Logic to determine is_sent
                normalized_sender = re.sub(r'[^0-9]', '', str(sender))
                normalized_owner = re.sub(r'[^0-9]', '', str(owner_val))
                is_sent = False
                if normalized_owner and normalized_owner in normalized_sender:
                    is_sent = True
                elif "owner" in sender.lower() or "proprietario" in sender.lower():
                    is_sent = True
                elif sender == owner_val:
                    is_sent = True
                
                body = clean_text(row[28])
                timestamp = str(row[37]) if row[37] is not None else ""
                
                att_name = row[45]
                att_info = str(att_name).strip() if att_name is not None and str(att_name).strip() != "" else None
                
                trans_raw = str(row[50]) if row[50] is not None else ""
                translation = ""
                if "Traduzione:" in trans_raw:
                    try:
                        translation = trans_raw.split("Traduzione:", 1)[1]
                        translation = clean_text(translation)
                    except: pass
                
                if body.strip().lower() in ["ok", "ok.", "ok!"] and len(translation) > 10:
                    translation = ""
                elif len(body) < 30 and len(translation) > (len(body) * 3 + 10):
                    translation = ""

                chats[chat_id]["messages"].append({
                    "sender": sender,
                    "is_sent": is_sent,
                    "body": body,
                    "time": timestamp,
                    "att": att_info,
                    "trans": translation
                })
            except Exception as e:
                # print(f"Row error: {e}")
                pass
        return chats

class WhatsAppParser(BaseParser):
    def parse(self, filepath):
        print("Using WhatsAppParser (Light)...")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb['Instant Messages']
        
        # Find headers
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: return {}
        
        header = rows[0]
        # Map columns
        col_map = {name: i for i, name in enumerate(header) if name}
        
        chats = {}
        
        for row in rows[1:]:
            try:
                def get_col(name):
                    return row[col_map[name]] if name in col_map and col_map[name] < len(row) else None

                from_val = get_col('From')
                to_val = get_col('To')
                body_val = get_col('Body')
                
                # Timestamp might be split or named differently
                time_val = get_col('Timestamp-Time')
                if time_val is None: time_val = get_col('Timestamp')
                
                if from_val is None and to_val is None: continue

                # Participants logic
                parts = set(re.findall(r'(\d+@s\.whatsapp\.net)', str(from_val) + " " + str(to_val)))
                if not parts: continue
                
                chat_id = "Chat " + "-".join(sorted([p.split('@')[0] for p in parts]))
                
                if chat_id not in chats:
                    chats[chat_id] = {
                        "id": chat_id,
                        "participants": " ".join(parts),
                        "owner": "Unknown",
                        "messages": []
                    }
                
                sender_match = re.search(r'(\d+@s\.whatsapp\.net)', str(from_val))
                sender = sender_match.group(1).split('@')[0] if sender_match else "Unknown"
                
                is_sent = False 
                
                body = clean_text(body_val)
                ts_str = str(time_val) if time_val is not None else ""
                
                chats[chat_id]["messages"].append({
                    "sender": sender,
                    "is_sent": is_sent,
                    "body": body,
                    "time": ts_str,
                    "att": None,
                    "trans": None
                })
                
            except Exception as e:
                pass
        
        return chats

# ==========================================
# RENDERERS
# ==========================================

class HTMLRenderer:
    def __init__(self, css, file_map, output_dir, style_mode="signal"):
        self.css = css
        self.file_map = file_map
        self.output_dir = output_dir
        self.style_mode = style_mode

    def render(self, chats, filename="index.html"):
        sidebar_html = ""
        chats_html = ""
        
        sorted_chat_ids = sorted(chats.keys())
        first = True
        
        for chat_id in sorted_chat_ids:
            data = chats[chat_id]
            msgs = data["messages"]

            # Parse Participants for Layout
            owner_name, owner_num, contact_name, contact_num = parse_participants_intelligent(data, chat_id)
            
            # Sidebar Item
            avatar_letter = contact_name[0].upper() if contact_name else "?"
            
            # Search Index Generation
            search_text = f"{contact_name} {contact_num} " + " ".join([f"{m['body']} {m['trans'] or ''}" for m in msgs])
            search_text = search_text.replace('"', '&quot;').lower()
            
            # SIDEBAR
            sidebar_html += f"""
            <div id="item-{chat_id}" class="chat-item" onclick="loadChat('{chat_id}')" data-search="{search_text}">
                <div class="sidebar-avatar">{avatar_letter}</div>
                <div style="flex:1;">
                    <div class="sidebar-name">{contact_name[:30]}</div>
                    <div class="sidebar-info">{len(msgs)} messaggi</div>
                </div>
            </div>
            """
            
            # Chat Area
            display_style = "flex" if first else "none"
            first = False
            
            msgs_html = ""
            current_date_str = None
            
            for idx, msg in enumerate(msgs):
                # Unique Message ID
                msg_id = f"msg-{chat_id}-{idx}"
                # Date Divider Logic
                try:
                    msg_date = msg['time'][:10] 
                    if msg_date != current_date_str and len(msg_date) >= 8 and ('/' in msg_date or '-' in msg_date):
                        current_date_str = msg_date
                        msgs_html += f'<div class="date-divider">{current_date_str}</div>'
                except: pass

                msg_class = "sent" if msg["is_sent"] else "received"
                
                sender_disp = msg["sender"]
                if msg["is_sent"]: 
                    sender_disp = "Tu"
                elif sender_disp in contact_num:
                    sender_disp = contact_name

                sender_div = f'<div class="sender-name">{sender_disp}</div>'
                
                # Attachment
                att_html = ""
                if msg["att"]:
                    basename = os.path.basename(msg["att"])
                    if basename in self.file_map:
                        rel_path = self.file_map[basename]
                        _, ext_raw = os.path.splitext(basename)
                        ext = ext_raw.replace('.', '').lower().strip()
                        
                        img_exts = ['jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp', 'svg', 'ico', 'tif', 'tiff', 'heic', 'heif', 'mms', 'mms_']
                        vid_exts = ['mp4', 'webm', 'ogg', 'mov', 'qt', 'avi', 'mkv', 'm4v', '3gp']
                        aud_exts = ['mp3', 'wav', 'aac', 'm4a', 'amr', 'opus', 'oga', 'flac']

                        if ext in img_exts:
                            att_html = f'''
                            <div class="attachment">
                                <img src="{rel_path}" onclick="window.open(this.src)" title="{basename}" onerror="this.style.display='none'; this.parentNode.innerHTML='<a href={rel_path} target=_blank>📷 {basename} (Caricamento fallito)</a>'">
                            </div>
                            '''
                        elif ext in vid_exts:
                            att_html = f'''
                            <div class="attachment">
                                <video controls style="max-width:300px; max-height:300px; border-radius:10px;">
                                    <source src="{rel_path}">
                                    <a href="{rel_path}">📹 {basename}</a>
                                </video>
                            </div>
                            '''
                        elif ext in aud_exts:
                            att_html = f'''
                            <div class="attachment">
                                <audio controls style="max-width:300px;">
                                    <source src="{rel_path}">
                                    <a href="{rel_path}">🎵 {basename}</a>
                                </audio>
                            </div>
                            '''
                        else:
                            att_html = f'<div class="attachment"><a href="{rel_path}" target="_blank">📄 {basename}</a></div>'
                
                trans_html = ""
                if msg["trans"]:
                    trans_html = f'<div class="translation">{msg["trans"]}</div>'
                
                msgs_html += f"""
                <div id="{msg_id}" class="message {msg_class}">
                    {sender_div}
                    <div class="msg-content">
                        {msg['body']}
                        {att_html}
                    </div>
                    {trans_html}
                    <div class="timestamp">{msg['time']}</div>
                </div>
                """
                
            # Header
            header_html = f"""
            <div class="chat-info-header">
                <div class="participant-card">
                    <div class="participant-avatar owner">T</div>
                    <div class="participant-details">
                        <span class="participant-name">{owner_name}</span>
                        <span class="participant-number">{owner_num}</span>
                    </div>
                </div>
                
                <div style="font-size:24px; color:#bbb">&#8596;</div>
                
                <div class="participant-card">
                    <div class="participant-avatar">{avatar_letter}</div>
                    <div class="participant-details">
                        <span class="participant-name">{contact_name}</span>
                        <span class="participant-number">{contact_num}</span>
                    </div>
                </div>
            </div>
            """
            
            chats_html += f"""
            <div id="chat-{chat_id}" class="chat-messages" style="display:{display_style};">
                {header_html}
                {msgs_html}
            </div>
            """
            
        # Decide Sidebar Header based on Style
        if self.style_mode == "whatsapp":
            sidebar_header_div = """
            <div class="chat-header-bar">
                <div class="participant-avatar owner" style="width:40px;height:40px;font-size:16px;border:0;">T</div>
                <span style="font-weight:600; margin-left:15px; font-size:16px;">WhatsApp Report</span>
            </div>
            <div class="search-box">
                <input type="text" class="search-input" placeholder="Cerca messaggi..." onkeyup="performSearch(this.value)">
            </div>
            """
        else:
            # Default Signal/Generic
            sidebar_header_div = f'<div class="signal-sidebar-header">Chats ({len(chats)})</div>'
            sidebar_header_div += f'''
            <div class="search-box">
                <input type="text" class="search-input" placeholder="Cerca messaggi..." onkeyup="performSearch(this.value)">
            </div>
            '''
            
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{self.css}</style>
            <script>
                function loadChat(chatId, msgId) {{
                    // Hide all chat views
                    document.querySelectorAll('.chat-messages').forEach(e => e.style.display = 'none');
                    
                    // Show target chat
                    var chatView = document.getElementById('chat-' + chatId);
                    if(chatView) chatView.style.display = 'flex';
                    
                    // Update Active Sidebar Item
                    document.querySelectorAll('.chat-item').forEach(e => e.classList.remove('active'));
                    var item = document.getElementById('item-' + chatId);
                    if(item) item.classList.add('active');
                    
                    // If msgId provided, scroll and highlight
                    if(msgId) {{
                        var msgEl = document.getElementById(msgId);
                        if(msgEl) {{
                            msgEl.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            msgEl.classList.remove('highlight-msg'); // reset
                            void msgEl.offsetWidth; // trigger reflow
                            msgEl.classList.add('highlight-msg');
                        }}
                    }}
                }}
                
                function performSearch(query) {{
                    var chatList = document.getElementById('chat-list');
                    var resultsContainer = document.getElementById('search-results');
                    var resultsContent = document.getElementById('search-results-content');
                    
                    if (!query || query.length < 2) {{
                        chatList.style.display = 'block';
                        resultsContainer.style.display = 'none';
                        return;
                    }}
                    
                    chatList.style.display = 'none';
                    resultsContainer.style.display = 'block';
                    resultsContent.innerHTML = '';
                    
                    var q = query.toLowerCase();
                    var count = 0;
                    var maxResults = 100;
                    
                    // Search Logic: Iterate all messages in DOM
                    // Optimization: We could use the data-search on chats to filter first?
                    // But user wants specific messages.
                    
                    var allMessages = document.querySelectorAll('.message');
                    for (var i = 0; i < allMessages.length; i++) {{
                        if(count >= maxResults) break;
                        var msg = allMessages[i];
                        
                        // Check content
                        var text = msg.innerText; // innerText includes sender, time, body, translation
                        if(text.toLowerCase().includes(q)) {{
                            count++;
                            
                            // Context
                            var chatContainer = msg.closest('.chat-messages');
                            var chatId = chatContainer.id.replace('chat-', '');
                            var msgId = msg.id;
                            
                            // Get Chat Name from sidebar item
                            var chatItem = document.getElementById('item-' + chatId);
                            var chatName = chatItem ? chatItem.querySelector('.sidebar-name').innerText : 'Chat ' + chatId;
                            
                        // Highlight snippet
                        var lowerText = text.toLowerCase();
                        var idx = lowerText.indexOf(q);
                        var start = Math.max(0, idx - 40);
                        var end = Math.min(text.length, idx + 60);
                        var rawSnippet = text.substring(start, end);
                        if (start > 0) rawSnippet = "..." + rawSnippet;
                        if (end < text.length) rawSnippet = rawSnippet + "...";
                        
                        // Highlight term visually
                        // Escape regex special chars
                        var qEscaped = q.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&');
                        var regex = new RegExp("(" + qEscaped + ")", "gi");
                        var styledSnippet = rawSnippet.replace(regex, '<span class="highlight-term">$1</span>');
                        
                        // Avatar styling
                        var avatarLetter = chatName.charAt(0).toUpperCase();
                        
                        var div = document.createElement('div');
                        div.className = 'search-result-item';
                        div.innerHTML = `
                            <div class="search-avatar">${{avatarLetter}}</div>
                            <div class="search-content">
                                <div class="search-result-sender">${{chatName}}</div>
                                <div class="search-result-preview">${{styledSnippet}}</div>
                            </div>
                        `;
                        
                        // Closure for click
                        (function(cid, mid) {{
                            div.onclick = function() {{ loadChat(cid, mid); }};
                        }})(chatId, msgId);
                        
                        resultsContent.appendChild(div);
                        }}
                    }}
                    
                    if (count === 0) {{
                        resultsContent.innerHTML = '<div style="padding:20px;text-align:center;color:#888">Nessun messaggio trovato.</div>';
                    }}
                }}
                
                function closeSearch() {{
                    document.querySelector('.search-input').value = '';
                    performSearch('');
                }}
            </script>
        </head>
        <body>
            <div class="container">
                <div class="sidebar">
                    {sidebar_header_div}
                    <div id="chat-list" style="flex:1; overflow-y:auto;">
                        {sidebar_html}
                    </div>
                    <div id="search-results" class="search-results-container">
                        <div class="search-header">
                            <span>RISULTATI RICERCA</span>
                            <span class="close-search" onclick="closeSearch()">CHIUDI</span>
                        </div>
                        <div id="search-results-content"></div>
                    </div>
                </div>
                <div class="main">
                    {chats_html}
                </div>
            </div>
        </body>
        </html>
        """
        
        path = os.path.join(self.output_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            f.write(full_html)
        print(f"Generated: {path}")

# ==========================================
# MAIN LOGIC
# ==========================================

def run_generation(input_file, style="signal", output_dir=None, format_type="auto"):
    input_path = os.path.abspath(input_file)
    base_dir = os.path.dirname(input_path)
    
    if output_dir:
        out_dir = output_dir
    else:
        out_dir = os.path.join(base_dir, "VisualReport")
        
    os.makedirs(out_dir, exist_ok=True)
    
    print(f"Input: {input_path}")
    print(f"Output: {out_dir}")
    print(f"Style: {style}")

    # 1. Determine Parser
    parser_impl = None
    if format_type == "cellebrite":
        parser_impl = CellebriteParser()
    elif format_type == "whatsapp":
        parser_impl = WhatsAppParser()
    else:
        # Auto detect
        try:
            wb = openpyxl.load_workbook(input_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            if "Chat" in sheet_names:
                print("Detected Cellebrite format (Chat sheet found)")
                parser_impl = CellebriteParser()
            elif "Instant Messages" in sheet_names:
                print("Detected WhatsApp format (Instant Messages sheet found)")
                parser_impl = WhatsAppParser()
            else:
                print("Could not detect format. Defaulting to Cellebrite.")
                parser_impl = CellebriteParser()
        except Exception as e:
            print(f"Error reading excel: {e}")
            return

    # 2. Parse
    chats = parser_impl.parse(input_path)
    print(f"Parsed {len(chats)} chats.")
    
    # 3. Scan attachments
    file_map = process_attachments(base_dir, out_dir)
    
    # 4. Render
    css = CSS_SIGNAL if style == "signal" else CSS_WHATSAPP
    renderer = HTMLRenderer(css, file_map, out_dir, style_mode=style)
    renderer.render(chats)
    print("Generation Complete.")
    return out_dir

# ==========================================
# GUI CLASSES
# ==========================================

class RedirectText:
    def __init__(self, text_ctrl):
        self.output = text_ctrl

    def write(self, string):
        self.output.insert(tk.END, string)
        self.output.see(tk.END)

    def flush(self):
        pass

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Chat Report Generator (Lite)")
        self.root.geometry("600x500")
        
        # Styles
        self.root.configure(bg="#f0f0f0")
        font_label = ("Segoe UI", 10)
        
        # 1. File Selection
        frame_file = tk.Frame(root, bg="#f0f0f0", pady=10)
        frame_file.pack(fill=tk.X, padx=20)
        
        tk.Label(frame_file, text="File Excel:", bg="#f0f0f0", font=font_label).pack(anchor="w")
        
        self.entry_file = tk.Entry(frame_file, width=50)
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=5)
        
        btn_browse = tk.Button(frame_file, text="Sfoglia...", command=self.browse_file, font=font_label)
        btn_browse.pack(side=tk.RIGHT, padx=5)
        
        # 2. Options
        frame_opts = tk.Frame(root, bg="#f0f0f0", pady=10)
        frame_opts.pack(fill=tk.X, padx=20)
        
        # Style
        tk.Label(frame_opts, text="Stile Grafico:", bg="#f0f0f0", font=font_label).grid(row=0, column=0, padx=5, sticky="w")
        self.var_style = tk.StringVar(value="signal")
        tk.Radiobutton(frame_opts, text="Signal (Blue)", variable=self.var_style, value="signal", bg="#f0f0f0").grid(row=0, column=1)
        tk.Radiobutton(frame_opts, text="WhatsApp (Green)", variable=self.var_style, value="whatsapp", bg="#f0f0f0").grid(row=0, column=2)
        
        # 3. Action
        frame_action = tk.Frame(root, bg="#f0f0f0", pady=20)
        frame_action.pack(fill=tk.X, padx=20)
        
        btn_run = tk.Button(frame_action, text="GENERA REPORT", command=self.start_generation, 
                            bg="#2c6bed", fg="white", font=("Segoe UI", 12, "bold"), height=2)
        btn_run.pack(fill=tk.X, pady=(0, 10))
        
        self.btn_open = tk.Button(frame_action, text="APRI CARTELLA OUTPUT", command=self.open_output, 
                                  state=tk.DISABLED, font=("Segoe UI", 10))
        self.btn_open.pack(fill=tk.X)
        
        # 4. Log Output
        tk.Label(root, text="Log:", bg="#f0f0f0", font=font_label).pack(anchor="w", padx=20)
        self.text_log = scrolledtext.ScrolledText(root, height=15)
        self.text_log.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Redirect stdout
        sys.stdout = RedirectText(self.text_log)
        sys.stderr = RedirectText(self.text_log)
        
        self.last_output_dir = None

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if filename:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, filename)

    def start_generation(self):
        input_path = self.entry_file.get()
        if not input_path or not os.path.exists(input_path):
            messagebox.showwarning("Attenzione", "Seleziona un file Excel valido.")
            return
            
        style = self.var_style.get()
        
        # Disable button
        self.btn_open.config(state=tk.DISABLED)
        
        t = threading.Thread(target=self.run_process, args=(input_path, style))
        t.start()
        
    def run_process(self, input_path, style):
        print("--- Inizio Elaborazione ---")
        try:
            # Determine output automatically (same dir)
            out_dir = run_generation(input_path, style=style, output_dir=None, format_type="auto")
            self.last_output_dir = out_dir
            
            self.root.after(0, lambda: self.finish_success())
        except Exception as e:
            error_msg = str(e)
            print(f"ERRORE: {error_msg}")
            self.root.after(0, lambda: messagebox.showerror("Errore", f"Si è verificato un errore:\n{error_msg}"))
        finally:
            print("--- Fine Elaborazione ---")

    def finish_success(self):
        messagebox.showinfo("Successo", "Report generato con successo!")
        self.btn_open.config(state=tk.NORMAL)
        
    def open_output(self):
        if self.last_output_dir and os.path.exists(self.last_output_dir):
            os.startfile(self.last_output_dir)
        else:
            messagebox.showwarning("Errore", "Cartella non trovata.")

# ==========================================
# ENTRY POINT
# ==========================================

def main():
    # If args passed, run CLI mode
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="Convert Excel Chat export to HTML")
        parser.add_argument("input_file", help="Path to Excel file")
        parser.add_argument("--style", choices=["signal", "whatsapp"], default="signal", help="Output visual style")
        parser.add_argument("--output", help="Output directory")
        parser.add_argument("--format", choices=["auto", "cellebrite", "whatsapp"], default="auto", help="Input format")
        
        args = parser.parse_args()
        run_generation(args.input_file, args.style, args.output, args.format)
    else:
        # GUI Mode
        root = tk.Tk()
        app = App(root)
        root.mainloop()

if __name__ == "__main__":
    main()
