import openpyxl
import re
from datetime import datetime

FILE_PATH = r"C:\Users\Administrator\Desktop\Sviluppo\Chat Insayets\2026-01-07.11-09-12\Rapporto motorola\Rapporto motorola_2026-01-07_Rapporto.xlsx"

def normalize_ws(s):
    return " ".join(str(s).split()).strip()

def analyze():
    print(f"Loading {FILE_PATH}...")
    try:
        wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    except Exception as e:
        print(f"Error loading file: {e}")
        return

    # 1. Inspect 'Chat' Sheet
    if 'Chat' not in wb.sheetnames:
        print("'Chat' sheet not found!")
        return
    
    sheet_chat = wb['Chat']
    rows_chat = list(sheet_chat.iter_rows(values_only=True))
    
    # Simple header detection for Chat (assuming row 1 based on previous logic, but let's print row 0-4)
    print("\n--- 'Chat' Sheet Header Analysis ---")
    header_chat = None
    chat_start_idx = 0
    chat_col_map = {}
    
    for i in range(min(5, len(rows_chat))):
        print(f"Row {i}: {rows_chat[i]}")
        row_str = [str(x).lower() for x in rows_chat[i] if x]
        if "timestamp" in row_str and "corpo" in row_str:
            header_chat = rows_chat[i]
            chat_start_idx = i + 1
            chat_col_map = {str(n).lower(): idx for idx, n in enumerate(header_chat) if n}
            print(f"-> Detected Header at Row {i}")
            break
            
    if not header_chat:
        print("Could not detect 'Chat' header.")
        return

    # 2. Inspect 'Messaggi istantanei' Sheet
    # Logic copied from WhatsAppParser to find the right sheet
    sheet_inst_name = None
    for name in wb.sheetnames:
        if "instant" in name.lower() or "messaggi" in name.lower():
            sheet_inst_name = name
            break
            
    if not sheet_inst_name:
        print("'Messaggi istantanei' sheet not found!")
        return

    sheet_inst = wb[sheet_inst_name]
    rows_inst = list(sheet_inst.iter_rows(values_only=True))
    
    print(f"\n--- '{sheet_inst_name}' Sheet Header Analysis ---")
    header_inst = None
    inst_start_idx = 0
    inst_col_map = {}
    
    for i in range(min(10, len(rows_inst))):
        row_str = [str(x).lower() for x in rows_inst[i] if x]
        if "informazioni sul file di origine" in row_str or "source file information" in row_str:
             header_inst = rows_inst[i]
             inst_start_idx = i + 1
             inst_col_map = {str(n).lower(): idx for idx, n in enumerate(header_inst) if n}
             print(f"-> Detected Header at Row {i}")
             break
             
    if not header_inst:
        print("Could not detect 'Instant Messages' header.")
        return

    # COLUMNS TO COMPARE
    # Chat: Timestamp, Body, Attachment
    # Instant: Timestamp, Body, Source Info
    
    c_time_idx = -1
    for k, v in chat_col_map.items():
        if "timestamp" in k: c_time_idx = v; break
        
    c_att_idx = chat_col_map.get("allegato", -1)
    if c_att_idx == -1: c_att_idx = chat_col_map.get("attachment", -1)
    
    c_body_idx = chat_col_map.get("corpo", -1)
    if c_body_idx == -1: c_body_idx = chat_col_map.get("body", -1)
    
    print(f"Chat Indices: Time={c_time_idx}, Att={c_att_idx}, Body={c_body_idx}")

    i_time_idx = -1
    for k, v in inst_col_map.items():
        if "timestamp" in k and "data" not in k and "ora" not in k: i_time_idx = v; break # Prefer combined
    if i_time_idx == -1: # Try separate? For now assume there's a main one or one of them
         for k, v in inst_col_map.items():
            if "timestamp" in k: i_time_idx = v; break
            
    i_src_idx = inst_col_map.get("informazioni sul file di origine", -1)
    i_body_idx = inst_col_map.get("corpo", -1)
    if i_body_idx == -1: i_body_idx = inst_col_map.get("body", -1)

    print(f"Instant Indices: Time={i_time_idx}, Source={i_src_idx}, Body={i_body_idx}")

    # COLLECT CHAT DATA FOR LOOKUP
    # Store as dictionary keyed by Timestamp (normalized) -> List of {Body, Attachment}
    # Using timestamp string directly might be flaky, but let's try.
    
    chat_lookup = {}
    
    count_valid_chat = 0
    for r in rows_chat[chat_start_idx:]:
        if not r: continue
        try:
            ts = str(r[c_time_idx]).strip()
            att = str(r[c_att_idx]).strip() if c_att_idx != -1 and r[c_att_idx] else ""
            body = normalize_ws(r[c_body_idx]) if c_body_idx != -1 and r[c_body_idx] else ""
            
            if ts not in chat_lookup: chat_lookup[ts] = []
            chat_lookup[ts].append({"att": att, "body": body})
            count_valid_chat += 1
        except IndexError: pass
        
    print(f"Indexed {count_valid_chat} rows from 'Chat'. Unique Timestamps: {len(chat_lookup)}")

    # CHECK MATCHES
    print("\n--- Checking Matches for Instant Messages with Attachments ---")
    matches = 0
    failures = 0
    scanned = 0
    
    for r in rows_inst[inst_start_idx:]:
        if not r: continue
        try:
            src_info = str(r[i_src_idx]) if i_src_idx != -1 and r[i_src_idx] else ""
            
            # Only care about rows with "part..." in source info, as per user description
            if "part" in src_info and ".mms" in src_info:
                scanned += 1
                ts = str(r[i_time_idx]).strip()
                body = normalize_ws(r[i_body_idx]) if i_body_idx != -1 and r[i_body_idx] else ""
                
                # Try to find in Chat
                candidates = chat_lookup.get(ts, [])
                
                # Heuristic: Match Body (if exists) or just take likely match if only 1
                found = None
                for c in candidates:
                    # If instant message has no body, ignore body match? 
                    # Signal often puts "part..." in source but logic says:
                    # Instant message body might be empty or specific string.
                    if body and c["body"] == body:
                        found = c
                        break
                    if not body and not c["body"]: # Both empty body
                        found = c 
                        break
                
                # If only 1 candidate and no body conflict, take it
                if not found and len(candidates) == 1:
                    found = candidates[0]
                
                if found:
                    print(f"[MATCH] Time: {ts}")
                    print(f"  Inst Source: ...{src_info[-50:]}")
                    print(f"  Chat Att:    {found['att']}")
                    print(f"  Body Match:  {body == found['body']}")
                    if found['att']: matches += 1
                    else: 
                        print("  (Chat row found but no Attachment listed)")
                else:
                    print(f"[FAIL] Time: {ts} | Candidates: {len(candidates)}")
                    failures += 1
                    
                if scanned >= 10: break # Only check first 10 relevant items
        except IndexError: pass

    print(f"\nAnalysis Complete. Scanned: {scanned}, Matches with Att: {matches}, Failures: {failures}")

if __name__ == "__main__":
    analyze()
