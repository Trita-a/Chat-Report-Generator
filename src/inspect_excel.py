import openpyxl
import sys
import os

def inspect(filepath):
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return

    print(f"--- Inspecting: {os.path.basename(filepath)} ---")
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        print(f"Sheets found: {wb.sheetnames}")
        
        target_sheets = [s for s in wb.sheetnames if "instant" in s.lower() or "messaggi" in s.lower() or "chat" in s.lower()]
        
        for sheet_name in target_sheets:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(max_row=5, values_only=True))
            
            for i, row in enumerate(rows):
                # Print non-none values for clarity
                print(f"Row {i}: {row}")
                
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        inspect(sys.argv[1])
    else:
        print("Usage: python src/inspect_excel.py <path_to_excel>")
