import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import sys
import os
import re
import argparse
import shutil
from datetime import datetime
import openpyxl
import json

# ==========================================
# CSS STYLES
# ==========================================

CSS_SIGNAL = """
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Oxygen, Ubuntu, Cantarell, "Open Sans", "Helvetica Neue", sans-serif; background-color: #f7f7f7; margin: 0; padding: 0; color: #1b1b1b; }
.container { display: flex; height: 100vh; }
.sidebar { width: 320px; flex-shrink: 0; background-color: white; border-right: 1px solid #e6e6e6; overflow-y: auto; }
.main { flex: 1; display: flex; flex-direction: column; background-color: #ffffff; }
.chat-messages { flex: 1; overflow-y: auto; padding: 20px 40px; display: flex; flex-direction: column; }
.message { max-width: 60%; margin-bottom: 8px; padding: 10px 14px; border-radius: 16px; position: relative; font-size: 15px; line-height: 1.4; word-wrap: break-word; }
.message.received { align-self: flex-start; background-color: #f2f2f2; color: #1b1b1b; border-bottom-left-radius: 4px; margin-right: auto; }
.message.sent { align-self: flex-end; background-color: #2c6bed; color: white; border-bottom-right-radius: 4px; margin-left: auto; }
.sender-name { font-size: 12px; font-weight: bold; margin-bottom: 2px; color: #666; }
.message.sent .sender-name { display: none; }
.timestamp { font-size: 10px; opacity: 0.6; float: right; margin-left: 8px; margin-top: 4px; }
.message.sent .timestamp { color: rgba(255,255,255,0.8); }
.message.has-attachment { padding: 5px; }
.message.has-attachment .attachment img { max-width: 200px; border-radius: 6px; margin: 0; }

/* SIDEBAR STYLES */
.signal-sidebar-header { font-size: 20px; font-weight: bold; padding: 24px 20px 16px 20px; border-bottom: 1px solid #e6e6e6; color: #1b1b1b; }
.chat-item { height: 72px; padding: 0 20px; border-bottom: 1px solid #f5f5f5; cursor: pointer; display: flex; flex-direction: row; align-items: center; transition: background-color 0.1s; }
.chat-item:hover { background-color: #f5f5f5; }
.chat-item.active { background-color: #d1e1fb; border-left: 4px solid #2c6bed; }
.sidebar-avatar { width: 52px; height: 52px; font-size: 20px; margin-right: 16px; margin-bottom: 0; display: flex; align-items: center; justify-content: center; background-color: #ddd; border-radius: 50%; color: #333; font-weight: bold; }
.sidebar-name { font-weight: 600; font-size: 15px; color: #1b1b1b; }
.sidebar-info { font-size: 13px; color: #666; margin-top: 3px; }

/* SHARED / OTHER */
.chat-info-header { background-color: #f7f7f7; padding: 15px; border-bottom: 1px solid #ddd; display: flex; justify-content: center; align-items: center; gap: 40px; margin: 10px 20px 20px 20px; border-radius: 8px; }
.participant-avatar { width: 60px; height: 60px; border-radius: 50%; background-color: #ddd; color: #333; display: flex; align-items: center; justify-content: center; font-size: 24px; font-weight: bold; margin-bottom: 8px; }
.participant-card { display: flex; flex-direction: column; align-items: center; text-align: center; }
.participant-name { font-weight: 700; font-size: 16px; color: #1b1b1b; }
.participant-number { font-size: 13px; color: #666; margin-top: 2px; }
.translation { margin-top: 8px; padding: 8px 10px; border-left: 3px solid #fecb00; font-style: normal; font-size: 13.5px; background-color: rgba(255,255,255,0.95); color: #333; border-radius: 6px; display: block; clear: both; margin-bottom: 2px; }
.date-divider { align-self: center; color: #555; font-size: 13px; margin: 15px 0; font-weight: 500; text-align: center; }
.message.sent .translation { background-color: rgba(255,255,255,0.15); color: white; border-left-color: rgba(255,255,255,0.5); }
.attachment img { max-width: 200px; border-radius: 12px; margin-top: 5px; cursor: pointer; }
.attachment video { max-width: 250px; border-radius: 12px; margin-top: 5px; }
.attachment audio { max-width: 250px; margin-top: 5px; }

/* Signal Audio Message Visual */
.audio-message { display: flex; align-items: center; gap: 12px; padding: 10px 14px; background: linear-gradient(135deg, #e8eaed 0%, #f5f5f5 100%); border-radius: 20px; width: 100%; box-sizing: border-box; }
.message.sent .audio-message { background: linear-gradient(135deg, #1a56db 0%, #2c6bed 100%); }
.audio-play-btn { width: 40px; height: 40px; border-radius: 50%; background: #2c6bed; display: flex; align-items: center; justify-content: center; cursor: pointer; flex-shrink: 0; }
.message.sent .audio-play-btn { background: rgba(255,255,255,0.25); }
.audio-play-btn::after { content: ""; width: 0; height: 0; border-left: 12px solid white; border-top: 7px solid transparent; border-bottom: 7px solid transparent; margin-left: 3px; }
.audio-play-btn.playing::after { border: none; width: 4px; height: 14px; background: white; box-shadow: 6px 0 0 white; margin-left: -5px; }
.audio-waveform-container { flex: 1; display: flex; flex-direction: column; gap: 4px; min-width: 0; }
.audio-waveform { display: flex; align-items: center; gap: 1px; height: 24px; }
.audio-waveform span { width: 2px; background: #b0c4de; border-radius: 1px; }
.audio-waveform span.played { background: #2c6bed; }
.message.sent .audio-waveform span { background: rgba(255,255,255,0.4); }
.message.sent .audio-waveform span.played { background: rgba(255,255,255,0.9); }
.audio-waveform span:nth-child(odd) { height: 60%; } .audio-waveform span:nth-child(even) { height: 100%; } .audio-waveform span:nth-child(3n) { height: 80%; } .audio-waveform span:nth-child(4n) { height: 45%; }
.audio-duration { font-size: 11px; color: #666; font-weight: 500; }
.message.sent .audio-duration { color: rgba(255,255,255,0.8); }
.media-btn { cursor: pointer; padding: 6px 12px; border-radius: 6px; background: #e0e0e0; margin-left: 20px; font-weight: 500; font-size: 13px; display: flex; align-items: center; gap: 6px; border: 1px solid #ccc; }
.media-btn:hover { background: #d0d0d0; }
.media-icon-btn { cursor: pointer; margin-left: auto; padding: 8px; border-radius: 50%; background: transparent; color: #666; transition: all 0.2s; display: flex; align-items: center; justify-content: center; }
.media-icon-btn:hover { background: #e8e8e8; color: #333; }
.audio-label { font-size: 12px; color: #666; margin-top: 6px; font-style: italic; }
.message.sent .audio-label { color: rgba(255,255,255,0.7); }
.search-box { padding: 10px 15px; background-color: #f7f7f7; border-bottom: 1px solid #e6e6e6; }
.search-input { width: 100%; padding: 8px 12px; border-radius: 6px; border: 1px solid #ddd; background-color: #e9e9e9; font-size: 14px; box-sizing: border-box; outline: none; }
.search-input:focus { background-color: white; border-color: #2c6bed; }

/* ADVANCED SEARCH - PREMIUM */
.search-results-container { display: none; flex: 1; overflow-y: auto; background-color: #fff; }
.search-header { padding: 16px 20px; font-weight: 700; font-size: 14px; color: #1b1b1b; background: #f7f7f7; border-bottom: 1px solid #e6e6e6; display: flex; justify-content: space-between; align-items: center; position: sticky; top: 0; z-index: 5; }
.close-search { cursor: pointer; color: #2c6bed; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; padding: 4px 8px; border-radius: 4px; transition: background 0.2s; }
.close-search:hover { background-color: rgba(44, 107, 237, 0.1); }
.search-result-item { padding: 12px 20px; border-bottom: 1px solid #f5f5f5; cursor: pointer; transition: background 0.2s; display: flex; align-items: center; }
.search-result-item:hover { background-color: #f0f2f5; }
.search-avatar { width: 42px; height: 42px; border-radius: 50%; background-color: #e0e0e0; color: #555; display: flex; align-items: center; justify-content: center; font-weight: bold; font-size: 16px; margin-right: 14px; flex-shrink: 0; }
.search-content { flex: 1; min-width: 0; }
.search-result-sender { font-size: 14px; font-weight: 600; color: #1b1b1b; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.search-result-preview { font-size: 13px; color: #666; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; line-height: 1.4; }
.highlight-term { background-color: rgba(255, 235, 59, 0.5); border-radius: 2px; color: #000; font-weight: 500; padding: 0 2px; }
.highlight-msg { position: relative; animation: flashMsg 2s ease-out; }
.highlight-msg::after { content: ""; position: absolute; top: 0; left: 0; right: 0; bottom: 0; background-color: rgba(255, 235, 59, 0.2); border-radius: inherit; pointer-events: none; opacity: 0; animation: fadeOverlay 3s forwards; }
@keyframes flashMsg { 0% { transform: scale(1.02); box-shadow: 0 4px 12px rgba(0,0,0,0.1); z-index: 10; } 100% { transform: scale(1); box-shadow: none; z-index: 1; } }
@keyframes fadeOverlay { 0% { opacity: 1; } 100% { opacity: 0; } }
"""

CSS_WHATSAPP = """
body { font-family: "Segoe UI", Roboto, Helvetica, Arial, sans-serif; background-color: #d1d7db; margin: 0; padding: 0; color: #111b21; }
.container { display: flex; height: 100vh; }
.sidebar { width: 350px; flex-shrink: 0; background-color: white; border-right: 1px solid #e9edef; overflow-y: auto; }
.main { flex: 1; display: flex; flex-direction: column; background-color: #efe7dd; background-image: url("https://user-images.githubusercontent.com/15075759/28719144-86dc0f70-73b1-11e7-911d-60d70fcded21.png"); }
.chat-messages { flex: 1; overflow-y: auto; padding: 20px 60px; display: flex; flex-direction: column; }
.message { max-width: 65%; margin-bottom: 8px; padding: 6px 7px 8px 9px; border-radius: 7.5px; position: relative; font-size: 14.2px; line-height: 19px; box-shadow: 0 1px 0.5px rgba(11,20,26,0.13); word-wrap: break-word; }
.message.received { align-self: flex-start; background-color: #ffffff; color: #111b21; border-top-left-radius: 0; margin-right: auto; }
.message.sent { align-self: flex-end; background-color: #d9fdd3; color: #111b21; border-top-right-radius: 0; margin-left: auto; }
.timestamp { font-size: 11px; float: right; margin-left: 10px; margin-top: 5px; color: #667781; }
.chat-item { height: 72px; display: flex; align-items: center; padding: 0 15px; border-bottom: 1px solid #f0f2f5; cursor: pointer; flex-direction: row; }
.chat-item:hover { background-color: #f5f6f6; }
.chat-item.active { background-color: #ebebeb; border-left: 4px solid #00a884; }
.chat-info-header { background-color: rgba(255,255,255,0.92); padding: 15px; border-radius: 12px; margin: 20px auto; max-width: 600px; display: flex; justify-content: center; align-items: center; gap: 30px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }
.chat-header-bar { background-color: #f0f2f5; padding: 10px 16px; border-left: 1px solid #d1d7db; display: flex; align-items: center; height: 60px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); z-index: 10; }
.participant-avatar { width: 60px; height: 60px; border-radius: 50%; background-color: #e9edef; color: #fff; display: flex; align-items: center; justify-content: center; font-size: 24px; font-weight: bold; margin-bottom: 8px; border: 3px solid white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
.participant-avatar.owner { background-color: #00a884; }
.participant-avatar.contact { background-color: #555; }
.participant-card { display: flex; flex-direction: column; align-items: center; text-align: center; }
.participant-name { font-weight: 600; font-size: 15px; color: #111b21; }
.participant-number { font-size: 12px; color: #667781; }
.sidebar-avatar { width: 45px; height: 45px; font-size: 18px; margin-right: 15px; margin-bottom: 0; display: flex; align-items: center; justify-content: center; background-color: #e9edef; border-radius: 50%; color: #fff; font-weight: bold; }
.sidebar-name { font-weight: 600; font-size: 15px; color: #111b21; }
.sidebar-info { font-size: 13px; color: #667781; }
.translation { margin-top: 6px; padding: 6px 8px; border-left: 3px solid #ffcc00; font-size: 13px; background-color: rgba(0,0,0,0.03); color: #333; border-radius: 4px; font-style: italic; }
.date-divider { text-align: center; align-self: center; margin: 10px 0; padding: 5px 12px; background-color: rgba(255,255,255,0.9); border-radius: 7.5px; box-shadow: 0 1px 0.5px rgba(11,20,26,0.13); color: #555; font-size: 12.5px; }
.attachment img { max-width: 200px; border-radius: 6px; margin-top: 4px; cursor: pointer; }
.attachment video { max-width: 200px; border-radius: 8px; margin-top: 5px; }
.attachment audio { max-width: 100%; margin-top: 5px; }
.message.received .sender-name { font-weight: bold; font-size: 13.5px; color: #d64937; margin-bottom: 4px; display: block; }
.message.sent .sender-name { font-weight: bold; font-size: 13.5px; color: #555; margin-bottom: 4px; display: block; opacity: 0.9; }
.message.has-attachment { padding: 5px; }
.message.has-attachment .attachment img { max-width: 200px; border-radius: 6px; margin: 0; }

/* Audio Transcription Visual - WhatsApp Style */
.audio-message { display: flex; align-items: center; gap: 10px; padding: 8px 12px; background: linear-gradient(135deg, #dcf8c6 0%, #d9fdd3 100%); border-radius: 20px; width: 100%; box-sizing: border-box; }
.message.received .audio-message { background: linear-gradient(135deg, #f0f0f0 0%, #ffffff 100%); }
.audio-play-btn { width: 36px; height: 36px; border-radius: 50%; background: #00a884; display: flex; align-items: center; justify-content: center; cursor: pointer; flex-shrink: 0; box-shadow: 0 1px 3px rgba(0,0,0,0.15); }
.message.received .audio-play-btn { background: #00a884; }
.audio-play-btn::after { content: ""; width: 0; height: 0; border-left: 10px solid white; border-top: 6px solid transparent; border-bottom: 6px solid transparent; margin-left: 2px; }
.audio-play-btn.playing::after { border: none; width: 3px; height: 12px; background: white; box-shadow: 5px 0 0 white; margin-left: -4px; }
.audio-waveform-container { flex: 1; display: flex; flex-direction: column; gap: 3px; min-width: 0; }
.audio-waveform { display: flex; align-items: center; gap: 1px; height: 22px; }
.audio-waveform span { width: 2px; background: #00a884; border-radius: 1px; opacity: 0.35; }
.audio-waveform span.played { opacity: 1.0; }
.audio-waveform span:nth-child(odd) { height: 60%; } .audio-waveform span:nth-child(even) { height: 100%; } .audio-waveform span:nth-child(3n) { height: 80%; } .audio-waveform span:nth-child(4n) { height: 45%; }
.audio-duration { font-size: 11px; color: #667781; font-weight: 500; }
.audio-label { font-size: 11px; color: #667781; margin-top: 4px; font-style: italic; }

.search-box { padding: 10px; background-color: #f0f2f5; border-bottom: 1px solid #d1d7db; }
.search-input { width: 100%; padding: 7px 12px; border-radius: 8px; border: none; background-color: #ffffff; font-size: 14px; box-sizing: border-box; height: 35px; outline: none; }

/* ADVANCED SEARCH - PREMIUM */
.search-results-container { display: none; flex: 1; overflow-y: auto; background-color: #fff; }
.search-header { padding: 16px 20px; font-weight: 700; font-size: 14px; color: #1b1b1b; background: #f7f7f7; border-bottom: 1px solid #e6e6e6; display: flex; justify-content: space-between; align-items: center; position: sticky; top: 0; z-index: 5; }
.close-search { cursor: pointer; color: #2c6bed; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; padding: 4px 8px; border-radius: 4px; transition: background 0.2s; }
.close-search:hover { background-color: rgba(44, 107, 237, 0.1); }
.search-result-item { padding: 12px 20px; border-bottom: 1px solid #f5f5f5; cursor: pointer; transition: background 0.2s; display: flex; align-items: center; }
.search-result-item:hover { background-color: #f0f2f5; }
.search-avatar { width: 42px; height: 42px; border-radius: 50%; background-color: #e0e0e0; color: #555; display: flex; align-items: center; justify-content: center; font-weight: bold; font-size: 16px; margin-right: 14px; flex-shrink: 0; }
.search-content { flex: 1; min-width: 0; }
.search-result-sender { font-size: 14px; font-weight: 600; color: #1b1b1b; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.search-result-preview { font-size: 13px; color: #666; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; line-height: 1.4; }
.highlight-term { background-color: rgba(255, 235, 59, 0.5); border-radius: 2px; color: #000; font-weight: 500; padding: 0 2px; }
.highlight-msg { position: relative; animation: flashMsg 2s ease-out; }
.highlight-msg::after { content: ""; position: absolute; top: 0; left: 0; right: 0; bottom: 0; background-color: rgba(255, 235, 59, 0.2); border-radius: inherit; pointer-events: none; opacity: 0; animation: fadeOverlay 3s forwards; }
@keyframes flashMsg { 0% { transform: scale(1.02); box-shadow: 0 4px 12px rgba(0,0,0,0.1); z-index: 10; } 100% { transform: scale(1); box-shadow: none; z-index: 1; } }
@keyframes fadeOverlay { 0% { opacity: 1; } 100% { opacity: 0; } }
"""

# ==========================================
# HELPERS
# ==========================================

def clean_text(text):
    if text is None:
        return ""
    
    text = str(text)
    if not text or text.lower() == 'nan':
         return ""

    patterns = [
        r'Etichette:.*?(?=\n|$)',
        r'Creato:.*?(?=\n|$)',
        r'Modificato:.*?(?=\n|$)',
        r'Descrizione:.*?(?=\n|$)',
        r'Generator:.*?(?=\n|$)',
        r'Tags:.*?(?=\n|$)',
        r'Created:.*?(?=\n|$)',
        r'Modified:.*?(?=\n|$)',
    ]
    cleaned = text
    for p in patterns:
        cleaned = re.sub(p, '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\n\s*\n', '\n', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\n\s*\n', '\n', cleaned).strip()
    cleaned = cleaned.replace("_x000d_", "")
    return cleaned

def parse_participants_intelligent(data, chat_id=""):
    """
    Intelligent logic to extract Owner and Contact from chat data.
    Returns: (owner_name, owner_number, contact_name, contact_number)
    """
    # 1. Get Owner from metadata if available (populated by WhatsAppParser)
    owner_name = data.get('owner', 'Unknown')
    owner_num = "" 
    if owner_name == "Unknown" or not owner_name:
         owner_name = "Proprietario"

    # Infer Owner from Messages if not known
    if owner_name == "Proprietario" or owner_name == "Tu":
        # Scan messages for is_sent = True
        msgs = data.get('messages', [])
        for m in msgs:
            if m.get('is_sent', False):
                possible_owner = m.get('sender')
                if possible_owner and possible_owner != "Tu" and possible_owner != "Proprietario":
                     owner_name = possible_owner
                     break
    
    # 2. Get Contact Name from participants field (already cleaned by Parser)
    contact_name = data.get('participants', '')
    contact_num = ""

    # If contact_name still contains "Sconosciuto" and we can't do better, keep it.
    if not contact_name or contact_name == "Unknown":
         contact_name = str(chat_id).replace("Chat ", "")

    # Fallback/Legacy Logic for Cellebrite Parser
    if "_x000d_" in contact_name or "\n" in contact_name:
        parts_raw = contact_name
        segments = re.split(r'_x000d_|\n', parts_raw)
        c_names = []
        for seg in segments:
            seg = seg.strip()
            if not seg: continue
            seg = re.sub(r'\(proprietario\)', '', seg, flags=re.IGNORECASE)
            seg = re.sub(r'\(owner\)', '', seg, flags=re.IGNORECASE).strip()
            if seg and "sconosciuto" not in seg.lower():
                 c_names.append(seg)
        if c_names:
             contact_name = " & ".join(c_names)

    # Clean Owner from Contact Name if present
    if owner_name and owner_name != "Proprietario" and owner_name != "Tu":
        if owner_name in contact_name and " & " in contact_name:
             # Basic replace: Remove "Alice & " or " & Alice"
             contact_name = contact_name.replace(f"{owner_name} & ", "").replace(f" & {owner_name}", "")
    
    # Final cleanup: Separate Number from Name if mixed "12345 Name"
    # This fixes "+39347... Name" appearing in sidebar
    if contact_name:
        # Regex for long number sequence
        match = re.search(r'(\+?\d{7,})', contact_name)
        if match:
            extracted_num = match.group(1)
            # Check if it's the whole string
            if len(extracted_num) == len(contact_name.replace(" ", "")):
                # It's just a number
                contact_num = extracted_num
                # keep contact_name as is or same? 
                # If just number, name is number.
            else:
                # It is "Number Name" or "Name Number"
                contact_num = extracted_num
                contact_name = contact_name.replace(extracted_num, "").strip(" -.,")
                if not contact_name: 
                    contact_name = contact_num # Fallback if name becomes empty

    # Final cleanup: Separate Number from Name for Owner (fixing user request)
    if owner_name and owner_num == "":
        match = re.search(r'(\+?\d{7,})', owner_name)
        if match:
            extracted_num = match.group(1)
            if len(extracted_num) == len(owner_name.replace(" ", "")):
                owner_num = extracted_num
            else:
                owner_num = extracted_num
                owner_name = owner_name.replace(extracted_num, "").strip(" -.,")
                if not owner_name: owner_name = owner_num

    return owner_name, owner_num, contact_name, contact_num

    return owner_name, owner_num, contact_name, contact_num

def process_attachments(source_dir, output_dir):
    """
    Scans source_dir for files. 
    Copies them to output_dir/attachments.
    Returns mapping {filename: 'attachments/filename'}
    """
    mapping = {}
    att_dir = os.path.join(output_dir, "attachments")
    os.makedirs(att_dir, exist_ok=True)
    
    print(f"Scanning and copying attachments from {source_dir}...")
    
    exclude_prefix = os.path.abspath(output_dir)
    
    for root, dirs, files in os.walk(source_dir):
        if os.path.abspath(root).startswith(exclude_prefix):
            continue
            
        for file in files:
            # Skip the report itself or script
            if file in ["index.html", "generate_chat_report.py", "ChatReportGenerator.py", "chat_report_gui.py", "ChatReportGenerator_Pandas.py"]:
                continue
                
            src_path = os.path.join(root, file)
            # Flatten structure: copy all unique filenames to attachments/
            dest_path = os.path.join(att_dir, file)
            
            try:
                if not os.path.exists(dest_path):
                    shutil.copy2(src_path, dest_path)
                
                # Rel path for HTML
                mapping[file] = f"attachments/{file}"
            except Exception as e:
                pass
                
    print(f"DEBUG: Scansione completata. Trovati {len(mapping)} file allegati.")
    return mapping

# ==========================================
# PARSERS (OpenPyXL Implementation)
# ==========================================

class BaseParser:
    def parse(self, filepath):
        raise NotImplementedError

class CellebriteParser(BaseParser):
    def parse(self, filepath):
        print("Using CellebriteParser (Light)...")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        sheet = None
        if 'Chat' in wb.sheetnames:
            sheet = wb['Chat']
        else:
            # Fallback to first
            sheet = wb.active
            
        chats = {}
        
        # Determine start row
        # Usually Cellebrite export has title then header. So row 1 is Title, Row 2 is Header, Row 3 is Data.
        # But indices in pandas were header=1 (so row 0 skip, row 1 header).
        # We will assume data starts from row 3 (index 3 in openpyxl) if checking pandas logic.
        # Let's start iterating from row 2 and see if it looks like data.
        
        # We'll skip first 2 rows.
        rows = list(sheet.iter_rows(values_only=True))
        
        # Heuristic: skip rows until we find one with an ID or looks like data
        # Pandas `header=1` means:
        # 0: "Report generated..."
        # 1: "ID", "Source", "Type"... (Header)
        # 2: 1, "WhatsApp", ... (Data)
        
        data_rows = rows[2:] # Slice off 0 and 1
        
        for idx, row in enumerate(data_rows):
            try:
                if not row or len(row) < 51: # Ensure row has minimal length
                    continue
                    
                # Index mapping (Pandas 0-based -> row[i])
                # 1: Chat ID
                chat_id = row[1]
                if chat_id is None: continue
                
                if chat_id not in chats:
                    chats[chat_id] = {
                        "id": chat_id,
                        "participants": str(row[9]) if row[9] is not None else "",
                        "owner": str(row[12]) if row[12] is not None else "",
                        "messages": []
                    }
                
                sender = str(row[21]) if row[21] is not None else ""
                owner_val = str(row[12]) if row[12] is not None else ""
                
                # Logic to determine is_sent
                normalized_sender = re.sub(r'[^0-9]', '', str(sender))
                normalized_owner = re.sub(r'[^0-9]', '', str(owner_val))
                is_sent = False
                if normalized_owner and normalized_owner in normalized_sender:
                    is_sent = True
                elif "owner" in sender.lower() or "proprietario" in sender.lower():
                    is_sent = True
                elif sender == owner_val:
                    is_sent = True
                
                body = clean_text(row[28])
                timestamp = str(row[37]) if row[37] is not None else ""
                
                att_name = row[45]
                att_info = str(att_name).strip() if att_name is not None and str(att_name).strip() != "" else None
                
                trans_raw = str(row[50]) if row[50] is not None else ""
                translation = ""
                if "Traduzione:" in trans_raw:
                    try:
                        translation = trans_raw.split("Traduzione:", 1)[1]
                        translation = clean_text(translation)
                    except: pass
                
                if body.strip().lower() in ["ok", "ok.", "ok!"] and len(translation) > 10:
                    translation = ""
                elif len(body) < 30 and len(translation) > (len(body) * 3 + 10):
                    translation = ""

                chats[chat_id]["messages"].append({
                    "sender": sender,
                    "is_sent": is_sent,
                    "body": body,
                    "time": timestamp,
                    "att": att_info,
                    "trans": translation
                })
            except Exception as e:
                # print(f"Row error: {e}")
                pass
        return chats

class WhatsAppParser(BaseParser):
    def parse(self, filepath, attachment_lookup=None):
        print(f"Using WhatsAppParser (Light)... Lookup size: {len(attachment_lookup) if attachment_lookup else 0}")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Try multiple variations of the sheet name
        sheet_name = 'Instant Messages'
        if 'Messaggi istantanei' in wb.sheetnames:
            sheet_name = 'Messaggi istantanei'
        elif 'Instant Messages' in wb.sheetnames:
            sheet_name = 'Instant Messages'
        else:
            # Fallback: try to find any sheet with "Instant" or "Messaggi"
            found = False
            for name in wb.sheetnames:
                if "instant" in name.lower() or "messaggi" in name.lower():
                    sheet_name = name
                    found = True
                    break
            
            if not found:
                 raise KeyError("Sheet 'Messaggi istantanei' or 'Instant Messages' not found.")

        print(f"Reading sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Find headers
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: 
            print("DEBUG: No rows found in sheet.")
            return {}
        
        if not rows: return {}
        
        # Search for header in first 10 rows
        header = None
        start_row_index = 0
        
        # Potential header keywords to look for
        keywords = ["Da", "A", "From", "To", "Corpo", "Body"]
        
        for i, row in enumerate(rows[:10]):
            row_vals = [str(c).lower() for c in row if c is not None]
            # Check if at least 2 keywords are in this row
            matches = sum(1 for k in keywords if k.lower() in row_vals)
            if matches >= 2:
                header = row
                start_row_index = i + 1
                print(f"Header found at row {i}: {header}")
                break
        
        if not header:
            # Fallback to row 0 if no header found (unlikely but safe)
            header = rows[0]
            start_row_index = 1
            print("Warning: Header not auto-detected, using first row.")

        # Map columns (handle both English and Italian)
        col_map = {name: i for i, name in enumerate(header) if name}
        
        chats = {}
        
        for row in rows[start_row_index:]:
            try:
                def get_col(names):
                    if isinstance(names, str): names = [names]
                    for n in names:
                        if n in col_map and col_map[n] < len(row):
                            return row[col_map[n]]
                    return None

                from_val = get_col(['From', 'Da'])
                to_val = get_col(['To', 'A'])
                
                if from_val is None and to_val is None: continue

                body_val = get_col(['Body', 'Corpo'])
                
                # Timestamp might be split or named differently
                time_val = get_col(['Timestamp-Time', 'Timestamp-Ora', 'Timestamp'])
                
                # Tag/Label for translation
                tag_val = get_col(['Tag', 'Etichetta', 'Note investigative', 'Messaggio Note investigative'])

                if from_val is None and to_val is None: continue

                direction_val = get_col(['Direction', 'Orientamento', 'Type', 'Tipo'])
                
                # Source info for attachments fallback
                source_info_val = get_col(['Informazioni sul file di origine', 'Source File Information', 'Source file information'])

                # Participants logic
                sender_raw = str(from_val)
                receiver_raw = str(to_val)
                
                def clean_participant(p):
                    # 1. Handle "number@s.whatsapp.net Name" or any variant
                    # Extract number and name
                    
                    cleaned_num = ""
                    cleaned_name = p
                    
                    # Regex for ID pattern: digits followed by @s.whatsapp.net
                    # We also want to catch just @s.whatsapp.net if no number
                    m_id_num = re.search(r'(\d+)@s\.whatsapp\.net', p)
                    if m_id_num:
                        cleaned_num = m_id_num.group(1)
                    
                    # Remove ANY @s.whatsapp.net and preceding potential ID junk
                    # This cleans "393...@s.whatsapp.net" -> ""
                    cleaned_name = re.sub(r'\S*@s\.whatsapp\.net', ' ', cleaned_name)
                    
                    # 2. Remove " - Delivered:..." metadata (and others)
                    split_pat = r'(?:\s+-\s+|-)(?:Inviato|Sent|Letti|Read|Delivered|Consegnato):'
                    parts = re.split(split_pat, cleaned_name, flags=re.IGNORECASE)
                    cleaned_name = parts[0]
                    
                    # 3. Remove (proprietario)/(owner)
                    cleaned_name = re.sub(r'\((?:proprietario|owner|device owner)\)', '', cleaned_name, flags=re.IGNORECASE)
                    
                    # 4. Cleanup garbage
                    cleaned_name = cleaned_name.replace('_x000d_', ' ').replace('\u200e', '').replace('\u202c', '')
                    cleaned_name = cleaned_name.replace('\n', ' ').replace('\r', ' ')
                    cleaned_name = re.sub(r'\s+', ' ', cleaned_name).strip()
                    cleaned_name = cleaned_name.strip('.,&;- ')
                    
                    # 5. Re-assemble: "Name Number"
                    if cleaned_num:
                        # If name has become empty or just characters like "-", use number
                        if not cleaned_name or len(cleaned_name) < 2:
                            return cleaned_num
                        
                        # If the name ALREADY contains the number, don't duplicate
                        if cleaned_num in cleaned_name.replace(" ", ""):
                            return cleaned_name
                            
                        # If name ends with number, keep custom logic?
                        # User wants Name + Number
                        return f"{cleaned_name} {cleaned_num}"
                    
                    return cleaned_name

                p1_clean = clean_participant(sender_raw)
                p2_clean = clean_participant(receiver_raw)
                
                # Filter out "Sconosciuto" if possible
                def is_valid_name(n):
                    if not n: return False
                    n_low = n.lower()
                    if "sconosciuto" in n_low or "unknown" in n_low: return False
                    if n_low == "&": return False
                    return True

                # Determine who is the "Owner" (Me)
                is_p1_owner = "(proprietario)" in sender_raw.lower() or "(owner)" in sender_raw.lower()
                is_p2_owner = "(proprietario)" in receiver_raw.lower() or "(owner)" in receiver_raw.lower()
                
                # Chat ID Generation
                parts = set(re.findall(r'(\d+@s\.whatsapp\.net)', sender_raw + " " + receiver_raw))
                if parts:
                    chat_id = "Chat " + "-".join(sorted([p.split('@')[0] for p in parts]))
                else:
                    # Use cleaned names for ID
                    parts_list = sorted([p for p in [p1_clean, p2_clean] if is_valid_name(p)])
                    if not parts_list:
                        # Fallback to whatever we have if both are sconosciuto
                        parts_list = sorted([p for p in [p1_clean, p2_clean] if p])
                        if not parts_list: parts_list = ["Unknown"]
                    
                    chat_id = f"Chat {' & '.join(parts_list)}"
                    chat_id = re.sub(r'[^\w\s\-\.]', '', chat_id)[:100]

                # Sender Logic
                sender = p1_clean
                
                # Direction Logic MOVED UP to help identify Owner
                is_sent = False 
                # 1. Try explicit 'Direction' column
                if direction_val:
                    d_str = str(direction_val).lower()
                    if "uscita" in d_str or "outgoing" in d_str:
                        is_sent = True
                    elif "entrata" in d_str or "incoming" in d_str:
                        is_sent = False
                    # Else fall through
                
                # 2. Fallback to owner tag logic if direction ambiguous or not found
                if not is_sent: # Only check if not already confirmed sent
                     if is_p1_owner and p1_clean == sender: 
                        is_sent = True

                # IMPROVED OWNER INFERENCE
                # If we know direction, we know who is owner (roughly)
                # Outgoing -> Sender is Owner
                # Incoming -> Receiver is Owner (assuming dump is from Owner's device)
                if not is_p1_owner and not is_p2_owner:
                    if is_sent:
                        is_p1_owner = True # Sender is Owner
                    else:
                        is_p2_owner = True # Receiver is Owner

                # Determine nice title for THIS row
                row_title = ""
                valid_names = [p for p in [p1_clean, p2_clean] if is_valid_name(p)]
                
                if is_p1_owner and p2_clean:
                    row_title = p2_clean
                elif is_p2_owner and p1_clean:
                    row_title = p1_clean
                elif len(valid_names) == 1:
                    row_title = valid_names[0]
                elif len(valid_names) >= 2:
                    row_title = " & ".join(valid_names)
                else:
                    # Both invalid/sconosciuto
                    row_title = "Chat Sconosciuta"

                if chat_id not in chats:
                    chats[chat_id] = {
                        "id": chat_id,
                        "participants": row_title, 
                        "owner": "Unknown",
                        "messages": []
                    }
                else:
                    # Update title if current one is bad and new one is better
                    curr_title = chats[chat_id]["participants"]
                    
                    # Logic: If current title is "Name & Name" (Ambiguous) but now we have a definitive single partner, overwrite it!
                    is_ambiguous = " & " in curr_title
                    is_better = " & " not in row_title and row_title != "Chat Sconosciuta"
                    
                    if ("sconosciuta" in curr_title.lower() or "unknown" in curr_title.lower()) and \
                       ("sconosciuta" not in row_title.lower() and "unknown" not in row_title.lower()):
                            chats[chat_id]["participants"] = row_title
                    elif is_ambiguous and is_better:
                            chats[chat_id]["participants"] = row_title
                
                # Update Owner if found
                current_owner = chats[chat_id]["owner"]
                if current_owner == "Unknown" or current_owner == "Proprietario":
                    if is_p1_owner:
                        chats[chat_id]["owner"] = p1_clean
                    elif is_p2_owner:
                        chats[chat_id]["owner"] = p2_clean

                
                body = clean_text(body_val)

                # Call Log Formatting
                # Example: "Vlora (...) started a call. status: Missed type: audio call duration: 00:00:00 ..."
                if "started a call" in body and "status:" in body:
                    try:
                        # Extract details
                        call_type = "Chiamata"
                        if "video call" in body: call_type = "Videochiamata"
                        elif "audio call" in body: call_type = "Chiamata Audio"
                        
                        # Extract Duration first to help decide status
                        duration = ""
                        has_duration = False
                        
                        m_dur = re.search(r'duration:\s*(\d{2}:\d{2}:\d{2})', body)
                        if m_dur:
                            dur_str = m_dur.group(1)
                            if dur_str != "00:00:00":
                                duration = f" ({dur_str})"
                                has_duration = True
                        
                        # Status Logic
                        status = "Sconosciuto"
                        
                        # parse status strictly
                        st_raw = "unknown"
                        m_st = re.search(r'status:\s*(\w+)', body, flags=re.IGNORECASE)
                        if m_st: st_raw = m_st.group(1).lower()
                        
                        if "missed" in st_raw:
                            status = "Persa"
                        elif "answered" in st_raw:
                            if is_sent:
                                status = "Effettuata"
                            else:
                                status = "Ricevuta"
                        elif "incoming" in st_raw:
                            if has_duration:
                                status = "Ricevuta"
                            else:
                                status = "In Arrivo (Senza Risposta)" # Or just In Arrivo?
                        elif "outgoing" in st_raw:
                            if has_duration:
                                status = "Effettuata"
                            else:
                                status = "In Uscita (Senza Risposta)"
                        
                        # Reformat body
                        body = f"📞 {call_type} {status}{duration}"
                    except: pass # Keep original parsing fails
                ts_str = str(time_val) if time_val is not None else ""
                
                # Attachment Processing
                att_info = None
                source_info = str(source_info_val) if source_info_val else ""
                
                # Check for "part...mms" in source info (Signal/WhatsApp native dumps)
                if "part" in source_info and ".mms" in source_info:
                     # Attempt to extract part filename
                     try:
                         # Ex: .../app_parts/part123.mms/part123.mms.
                         # Logic: look for part<digits>.mms or similar
                         m = re.search(r'(part\d+\.mms)', source_info)
                         if m:
                             att_info = m.group(1) + "_" # Often has trailing underscore on disk
                         else:
                             att_info = "Attachment (unknown)"
                     except: pass
                
                # Check for .opus audio files (WhatsApp voice messages)
                # Check for .opus audio files (WhatsApp voice messages)
                if ".opus" in source_info.lower():
                    try:
                        # Extract the opus filename - try multiple patterns
                        # Pattern 1: UUID-style like "e78b07cb-a66d-43b3-aa75-766ac54d2b45.opus"
                        # Pattern 2: Any filename before .opus
                        m_opus = re.search(r'([a-zA-Z0-9_\-]+\.opus)', source_info, flags=re.IGNORECASE)
                        if m_opus:
                            att_info = m_opus.group(1)
                    except: pass
                
                # Check for Images/Videos if not yet found
                # Check for Images/Videos if not yet found
                if not att_info:
                    try:
                        # Common image/video extensions
                        exts = r"(?:jpg|jpeg|png|gif|webp|bmp|heic|heif|tif|tiff|mp4|webm|mov|avi|mkv|3gp|m4v)"
                        # Debug image detection
                        if any(x in source_info.lower() for x in ['.jpg', '.jpeg', '.png', '.mp4']):
                             print(f"DEBUG: Potential media checking: {source_info[:50]}...")
                        
                        # Regex for filename ending with extension
                        # Try to capture more permissive filenames (allow spaces, parens)
                        m_media = re.search(r'([^\\/:*?"<>|\r\n]+\.' + exts + r')', source_info, flags=re.IGNORECASE)
                        if m_media:
                            result = m_media.group(1).strip()
                            # If result looks like a path, take basename
                            if "/" in result or "\\" in result:
                                result = os.path.basename(result)
                            att_info = result
                            print(f"DEBUG: Extracted media: {att_info}")
                        else:
                            if any(x in source_info.lower() for x in ['.jpg', '.jpeg', '.png', '.mp4']):
                                print(f"DEBUG: Failed to extract media from: {source_info}")
                    except Exception as e:
                        print(f"DEBUG: Error extracting media: {e}")
                        pass
                
                # CROSS-REFERENCE LOOKUP
                if attachment_lookup and ts_str:
                     # Clean timestamp to match key
                     # Expect key to be: "20/05/2025 21:45:46" (roughly)
                     # Valid keys in map are strings.
                     # Remove (UTC...)
                     ts_clean = re.sub(r'\(UTC.*?\)', '', ts_str).strip()
                     
                     if ts_clean in attachment_lookup:
                         real_att = attachment_lookup[ts_clean]
                         if real_att:
                             att_info = real_att # OVERRIDE with real name
                
                # Translation extraction from Tag
                trans_raw = str(tag_val) if tag_val is not None else ""
                translation = ""
                
                # Filter out known non-translation tags
                ignore_tags = ["deleted", "read", "delivered", "sent", "cancellato", "letto", "consegnato", "inviato", "none", "nan"]
                
                # Case-insensitive checks using regex
                # This handles "Traduzione:", "traduzione:", "Translation:", "translation:"
                m_trans = re.search(r'(?:Traduzione|Translation):', trans_raw, flags=re.IGNORECASE)
                
                if m_trans:
                    try:
                        # Split using the regex match
                        translation = re.split(r'(?:Traduzione|Translation):', trans_raw, maxsplit=1, flags=re.IGNORECASE)[1]
                        
                        # Often followed by "Etichette:" or other fields
                        if "Etichette:" in translation:
                            translation = translation.split("Etichette:", 1)[0]
                        if "Tags:" in translation:
                            translation = translation.split("Tags:", 1)[0]
                        if "Description:" in translation:
                            translation = translation.split("Description:", 1)[0]
                        translation = clean_text(translation)
                    except: pass
                elif "Descrizione:" in trans_raw:
                    try:
                        translation = trans_raw.split("Descrizione:", 1)[1]
                        # Often followed by "Etichette:", "Creato:", "Modificato:"
                        if "Etichette:" in translation:
                            translation = translation.split("Etichette:", 1)[0]
                        if "Tags:" in translation:
                            translation = translation.split("Tags:", 1)[0]
                        if "Description:" in translation:
                            translation = translation.split("Description:", 1)[0]
                        if "Creato:" in translation:
                            translation = translation.split("Creato:", 1)[0]
                        if "Created:" in translation:
                            translation = translation.split("Created:", 1)[0]
                        translation = clean_text(translation)
                        if not translation.strip(): translation = ""
                    except: pass
                elif trans_raw.strip() and trans_raw.lower() not in ignore_tags:
                    # Fallback: Assume the whole tag is a translation if it's not a status tag
                    
                    # 1. Clean "Description:" or "Descrizione:" from START of string
                    # e.g. "Description: Hello" -> "Hello"
                    # e.g. "Description:" -> ""
                    clean_trans = re.sub(r'^(?:Description|Descrizione):\s*', '', trans_raw, flags=re.IGNORECASE)
                    
                    # 2. Exclude simple numbers or short codes if resulting string is too short
                    if len(clean_trans) > 2 and not clean_trans.isdigit():
                         translation = clean_text(clean_trans)
                         if not translation.strip(): translation = ""

                chats[chat_id]["messages"].append({
                    "sender": sender,
                    "is_sent": is_sent,
                    "body": body,
                    "time": ts_str,
                    "att": att_info,
                    "trans": translation
                })
                
            except Exception as e:
                # print(f"Row error: {e}")
                pass
        
        return chats

# ==========================================
# RENDERERS
# ==========================================

class HTMLRenderer:
    def __init__(self, css, file_map, output_dir, style_mode="signal"):
        self.css = css
        self.file_map = file_map
        self.output_dir = output_dir
        self.style_mode = style_mode

    def render(self, chats, filename="index.html"):
        sidebar_html = ""
        chats_html = ""
        all_media = {} # {chat_id: {images:[], videos:[], links:[]}}
        
        sorted_chat_ids = sorted(chats.keys())
        first = True
        
        for chat_id in sorted_chat_ids:
            data = chats[chat_id]
            msgs = data["messages"]

            # SORT MESSAGES CHRONOLOGICALLY
            def parse_ts(m):
                ts_str = m.get("time", "")
                if not ts_str: return datetime.min
                # Remove (UTC...)
                clean = re.sub(r'\(UTC.*?\)', '', ts_str).strip()
                try:
                    # Try common formats
                    # 20/05/2025 19:00:03
                    return datetime.strptime(clean, "%d/%m/%Y %H:%M:%S")
                except:
                    try:
                        return datetime.strptime(clean, "%d/%m/%Y %H:%M")
                    except:
                        try:
                            # Try just date? or different sep
                            return datetime.strptime(clean, "%Y-%m-%d %H:%M:%S")
                        except:
                            return datetime.min # Fail safe

            msgs.sort(key=parse_ts)

            # Parse Participants for Layout
            owner_name, owner_num, contact_name, contact_num = parse_participants_intelligent(data, chat_id)
            
            # Sidebar Item
            avatar_letter = contact_name[0].upper() if contact_name else "?"
            
            # Dynamic Owner Avatar
            owner_initial = owner_name[0].upper() if owner_name and owner_name != "Unknown" else "P"
            
            # Search Index Generation
            search_text = f"{contact_name} {contact_num} " + " ".join([f"{m['body']} {m['trans'] or ''}" for m in msgs])
            search_text = search_text.replace('"', '&quot;').lower()
            
            # SIDEBAR
            sidebar_html += f"""
            <div id="item-{chat_id}" class="chat-item" onclick="loadChat('{chat_id}')" data-search="{search_text}">
                <div class="sidebar-avatar">{avatar_letter}</div>
                <div style="flex:1;">
                    <div class="sidebar-name">{contact_name[:30]}</div>
                    <div class="sidebar-info">{len(msgs)} messaggi</div>
                </div>
            </div>
            """
            
            # Chat Area
            display_style = "flex" if first else "none"
            first = False
            
            msgs_html = ""
            current_date_str = None
            media_collector = {'images': [], 'videos': [], 'links': [], 'docs': []}
            
            for idx, msg in enumerate(msgs):
                # Unique Message ID
                msg_id = f"msg-{chat_id}-{idx}"
                
                # Media Collection
                # Images/Videos
                if msg["att"]:
                    bname = os.path.basename(msg["att"])
                    if bname in self.file_map:
                        rpath = self.file_map[bname]
                        _, ext_raw = os.path.splitext(bname)
                        ext = ext_raw.replace('.', '').lower().strip()
                        
                        img_exts_c = ['jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp', 'svg', 'heic']
                        vid_exts_c = ['mp4', 'webm', 'ogg', 'mov', 'avi', 'mkv']
                        doc_exts_c = ['pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'txt', 'rtf', 'csv', 'zip', 'rar']
                        
                        if ext in img_exts_c:
                            media_collector['images'].append({'src': rpath, 'name': bname, 'msg_id': msg_id})
                        elif ext in vid_exts_c:
                            media_collector['videos'].append({'src': rpath, 'name': bname, 'msg_id': msg_id})
                        elif ext in doc_exts_c:
                            media_collector['docs'].append({'src': rpath, 'name': bname, 'msg_id': msg_id, 'ext': ext})

                # Links
                try:
                    urls = re.findall(r'(https?://[^\s]+)', msg['body'])
                    for u in urls:
                         media_collector['links'].append({'url': u, 'msg_id': msg_id})
                except: pass

                # Date Divider Logic
                try:
                    msg_date = msg['time'][:10] 
                    if msg_date != current_date_str and len(msg_date) >= 8 and ('/' in msg_date or '-' in msg_date):
                        current_date_str = msg_date
                        msgs_html += f'<div class="date-divider">{current_date_str}</div>'
                except: pass

                msg_class = "sent" if msg["is_sent"] else "received"
                
                sender_disp = msg["sender"]
                if msg["is_sent"]: 
                    sender_disp = owner_name if owner_name and owner_name != "Unknown" else "Tu"
                else:
                    # Received: prefer contact name over number
                    sender_disp = contact_name if contact_name else sender_disp

                sender_div = f'<div class="sender-name">{sender_disp}</div>'
                
                # Attachment
                att_html = ""
                if msg["att"]:
                    basename = os.path.basename(msg["att"])
                    if basename in self.file_map:
                        rel_path = self.file_map[basename]
                        _, ext_raw = os.path.splitext(basename)
                        ext = ext_raw.replace('.', '').lower().strip()
                        
                        img_exts = ['jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp', 'svg', 'ico', 'tif', 'tiff', 'heic', 'heif', 'mms', 'mms_']
                        vid_exts = ['mp4', 'webm', 'ogg', 'mov', 'qt', 'avi', 'mkv', 'm4v', '3gp']
                        aud_exts = ['mp3', 'wav', 'aac', 'm4a', 'amr', 'opus', 'oga', 'flac']

                        if ext in img_exts:
                            att_html = f'''
                            <div class="attachment">
                                <img src="{rel_path}" onclick="openLightbox('{rel_path}', 'image', '{basename}')" title="{basename}" style="cursor:pointer;" onerror="this.style.display='none'; this.parentNode.innerHTML='<a href={rel_path} target=_blank>📷 {basename} (Caricamento fallito)</a>'">
                            </div>
                            '''
                        elif ext in vid_exts:
                            att_html = f'''
                            <div class="attachment video-attachment" onclick="openLightbox('{rel_path}', 'video', '{basename}')" style="cursor:pointer; position:relative;">
                                <video src="{rel_path}" style="max-width:300px; max-height:300px; border-radius:10px; pointer-events:none;" preload="metadata"></video>
                                <div class="video-play-overlay" style="position:absolute; top:50%; left:50%; transform:translate(-50%,-50%); background:rgba(0,0,0,0.6); color:white; width:50px; height:50px; border-radius:50%; display:flex; align-items:center; justify-content:center; font-size:20px;">▶</div>
                            </div>
                            '''
                        elif ext in aud_exts:
                            # Skip if handled by custom Opus player (Visual Waveform)
                            is_special_opus = (ext == 'opus' and msg.get("trans") and not str(msg.get("body", "")).strip())
                            
                            if not is_special_opus:
                                att_html = f'''
                                <div class="attachment">
                                    <audio controls style="max-width:300px;">
                                        <source src="{rel_path}">
                                        <a href="{rel_path}">🎵 {basename}</a>
                                    </audio>
                                </div>
                                '''
                        else:
                            att_html = f'<div class="attachment"><a href="{rel_path}" target="_blank">📄 {basename}</a></div>'
                
                trans_html = ""
                is_audio_transcription = False
                
                # Detect audio transcription: has translation and is likely audio
                # Case 1: Has translation, no body, and opus attachment -> playable audio with transcription
                # Case 2: Has translation, no body, no attachment -> visual-only transcription
                
                is_opus_audio = msg["att"] and ".opus" in str(msg["att"]).lower()
                has_trans_only = msg["trans"] and not msg["body"].strip()
                
                if has_trans_only and is_opus_audio:
                    # Playable audio with transcription
                    is_audio_transcription = True
                    
                    # Get audio file path
                    basename = os.path.basename(msg["att"])
                    
                    # Prepare Custom Audio UI
                    text_len = len(msg["trans"])
                    num_bars = max(20, min(150, int(text_len * 2.3)))
                    waveform_spans = "<span></span>" * num_bars
                    
                    audio_element_id = f"audio-{chat_id}-{idx}"
                    onclick_action = ""
                    hidden_audio = ""
                    dur_text = "Messaggio Vocale"
                    
                    if basename in self.file_map:
                        rel_path = self.file_map[basename]
                        # Playable
                        onclick_action = f"onclick=\"playAudio(this, '{audio_element_id}')\""
                        hidden_audio = f'<audio id="{audio_element_id}" src="{rel_path}" hidden></audio>'
                    else:
                        # Not found
                        dur_text += " (Non trovato)"
                        onclick_action = "style='opacity:0.5; cursor:default;'"
                        
                    audio_html = f'''
                    <div class="audio-message">
                        <div class="audio-play-btn" {onclick_action}></div>
                        <div class="audio-waveform-container">
                            <div class="audio-waveform">
                                {waveform_spans}
                            </div>
                            <div class="audio-duration">{dur_text}</div>
                        </div>
                        {hidden_audio}
                    </div>
                    '''
                    
                    trans_html = f'''
                    {audio_html}
                    <div class="audio-label">Trascrizione:</div>
                    <div class="translation">{msg["trans"]}</div>
                    '''
                elif has_trans_only and not msg["att"]:
                    # Visual-only transcription (no audio file available)
                    is_audio_transcription = True
                    
                    # Calculate waveform bars based on text length
                    text_len = len(msg["trans"])
                    num_bars = max(20, min(150, int(text_len * 2.3)))
                    waveform_spans = "<span></span>" * num_bars
                    
                    audio_html = f'''
                    <div class="audio-message">
                        <div class="audio-play-btn" style="opacity:0.5; cursor:default;"></div>
                        <div class="audio-waveform-container">
                             <div class="audio-waveform">
                                {waveform_spans}
                            </div>
                           <div class="audio-duration">Messaggio Vocale (file non disponibile)</div>
                        </div>
                    </div>
                    '''

                    trans_html = f'''
                    {audio_html}
                    <div class="audio-label">Trascrizione:</div>
                    <div class="translation">{msg["trans"]}</div>
                    '''
                    

                elif msg["trans"]:
                    trans_html = f'<div class="translation">{msg["trans"]}</div>'
                
                has_att_class = " has-attachment" if msg["att"] else ""
                msgs_html += f"""
                <div id="{msg_id}" class="message {msg_class}{has_att_class}">
                    {sender_div}
                    <div class="msg-content">
                        {msg['body']}
                        {att_html}
                    </div>
                    {trans_html}
                    <div class="timestamp">{msg['time']}</div>
                </div>
                """
                
            # Header
            owner_num_disp = owner_num if owner_num else "(Numero non presente)"
            contact_num_disp = contact_num if contact_num else "(Numero non presente)"

            header_html = f"""
            <div class="chat-info-header">
                <div class="participant-card">
                    <div class="participant-avatar owner">{owner_initial}</div>
                    <div class="participant-details">
                        <span class="participant-name">{owner_name}</span>
                        <span class="participant-number">{owner_num_disp}</span>
                    </div>
                </div>
                
                <div style="font-size:24px; color:#bbb">&#8596;</div>
                
                <div class="participant-card">
                    <div class="participant-avatar">{avatar_letter}</div>
                    <div class="participant-details">
                        <span class="participant-name">{contact_name}</span>
                        <span class="participant-number">{contact_num_disp}</span>
                    </div>
                </div>
                
                <div class="media-icon-btn" onclick="openGallery('chat-{chat_id}')" title="Media, Link e Documenti">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="3" y="3" width="18" height="18" rx="2" ry="2"/>
                        <circle cx="8.5" cy="8.5" r="1.5"/>
                        <polyline points="21,15 16,10 5,21"/>
                    </svg>
                </div>
            </div>
            """
            
            chats_html += f"""
            <div id="chat-{chat_id}" class="chat-messages" style="display:{display_style};">
                {header_html}
                {msgs_html}
            </div>
            """
            all_media[f"chat-{chat_id}"] = media_collector
            
        # Decide Sidebar Header based on Style
        if self.style_mode == "whatsapp":
            sidebar_header_div = """
            <div class="chat-header-bar">
                <div class="participant-avatar owner" style="width:40px;height:40px;font-size:16px;border:0;">T</div>
                <span style="font-weight:600; margin-left:15px; font-size:16px;">WhatsApp Report</span>
            </div>
            <div class="search-box">
                <input type="text" class="search-input" placeholder="Cerca messaggi..." onkeyup="performSearch(this.value)">
            </div>
            """
        else:
            # Default Signal/Generic
            sidebar_header_div = f'<div class="signal-sidebar-header">Chats ({len(chats)})</div>'
            sidebar_header_div += f'''
            <div class="search-box">
                <input type="text" class="search-input" placeholder="Cerca messaggi..." onkeyup="performSearch(this.value)">
            </div>
            '''
            
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{self.css}</style>
            <script>
                function loadChat(chatId, msgId) {{
                    // Hide all chat views
                    document.querySelectorAll('.chat-messages').forEach(e => e.style.display = 'none');
                    
                    // Show target chat
                    var chatView = document.getElementById('chat-' + chatId);
                    if(chatView) chatView.style.display = 'flex';
                    
                    // Update Active Sidebar Item
                    document.querySelectorAll('.chat-item').forEach(e => e.classList.remove('active'));
                    var item = document.getElementById('item-' + chatId);
                    if(item) item.classList.add('active');
                    
                    // If msgId provided, scroll and highlight
                    if(msgId) {{
                        var msgEl = document.getElementById(msgId);
                        if(msgEl) {{
                            msgEl.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            msgEl.classList.remove('highlight-msg'); // reset
                            void msgEl.offsetWidth; // trigger reflow
                            msgEl.classList.add('highlight-msg');
                        }}
                    }}
                }}
                
                function performSearch(query) {{
                    var chatList = document.getElementById('chat-list');
                    var resultsContainer = document.getElementById('search-results');
                    var resultsContent = document.getElementById('search-results-content');
                    
                    if (!query || query.length < 2) {{
                        chatList.style.display = 'block';
                        resultsContainer.style.display = 'none';
                        return;
                    }}
                    
                    chatList.style.display = 'none';
                    resultsContainer.style.display = 'block';
                    resultsContent.innerHTML = '';
                    
                    var q = query.toLowerCase();
                    var count = 0;
                    var maxResults = 100;
                    
                    // Search Logic: Iterate all messages in DOM
                    // Optimization: We could use the data-search on chats to filter first?
                    // But user wants specific messages.
                    
                    var allMessages = document.querySelectorAll('.message');
                    for (var i = 0; i < allMessages.length; i++) {{
                        if(count >= maxResults) break;
                        var msg = allMessages[i];
                        
                        // Check content
                        var text = msg.innerText; // innerText includes sender, time, body, translation
                        if(text.toLowerCase().includes(q)) {{
                            count++;
                            
                            // Context
                            var chatContainer = msg.closest('.chat-messages');
                            var chatId = chatContainer.id.replace('chat-', '');
                            var msgId = msg.id;
                            
                            // Get Chat Name from sidebar item
                            var chatItem = document.getElementById('item-' + chatId);
                            var chatName = chatItem ? chatItem.querySelector('.sidebar-name').innerText : 'Chat ' + chatId;
                            
                        // Highlight snippet
                        var lowerText = text.toLowerCase();
                        var idx = lowerText.indexOf(q);
                        var start = Math.max(0, idx - 40);
                        var end = Math.min(text.length, idx + 60);
                        var rawSnippet = text.substring(start, end);
                        if (start > 0) rawSnippet = "..." + rawSnippet;
                        if (end < text.length) rawSnippet = rawSnippet + "...";
                        
                        // Highlight term visually
                        // Escape regex special chars
                        var qEscaped = q.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&');
                        var regex = new RegExp("(" + qEscaped + ")", "gi");
                        var styledSnippet = rawSnippet.replace(regex, '<span class="highlight-term">$1</span>');
                        
                        // Avatar styling
                        var avatarLetter = chatName.charAt(0).toUpperCase();
                        
                        var div = document.createElement('div');
                        div.className = 'search-result-item';
                        div.innerHTML = `
                            <div class="search-avatar">${{avatarLetter}}</div>
                            <div class="search-content">
                                <div class="search-result-sender">${{chatName}}</div>
                                <div class="search-result-preview">${{styledSnippet}}</div>
                            </div>
                        `;
                        
                        // Closure for click
                        (function(cid, mid) {{
                            div.onclick = function() {{ loadChat(cid, mid); }};
                        }})(chatId, msgId);
                        
                        resultsContent.appendChild(div);
                        }}
                    }}
                    
                    if (count === 0) {{
                        resultsContent.innerHTML = '<div style="padding:20px;text-align:center;color:#888">Nessun messaggio trovato.</div>';
                    }}
                }}
                
                function closeSearch() {{
                    document.querySelector('.search-input').value = '';
                    performSearch('');
                }}
                
                function playAudio(btn, audioId) {{
                    var audio = document.getElementById(audioId);
                    if (!audio) return;
                    
                    // Attach progress listener
                    if (!audio.dataset.hasListener) {{
                        audio.addEventListener('timeupdate', function() {{
                             var pct = audio.currentTime / audio.duration;
                             var container = btn.nextElementSibling;
                             if(container && container.classList.contains('audio-waveform-container')) {{
                                 var waveform = container.querySelector('.audio-waveform');
                                 if(waveform) {{
                                     var spans = waveform.children;
                                     var total = spans.length;
                                     var activeCount = Math.ceil(total * pct);
                                     for(var i=0; i<total; i++) {{
                                         if(i <= activeCount) spans[i].classList.add('played');
                                         else spans[i].classList.remove('played');
                                     }}
                                 }}
                             }}
                        }});
                        audio.dataset.hasListener = "true";
                    }}
                    
                    if (audio.paused) {{
                        try {{
                           document.querySelectorAll('audio').forEach(function(a){{ if(a.id!==audioId) {{ a.pause(); }} }});
                        }} catch(e){{}}
                        
                        audio.play();
                        btn.classList.add('playing');
                    }} else {{
                        audio.pause();
                        btn.classList.remove('playing');
                    }}
                    
                    audio.onended = function() {{
                        btn.classList.remove('playing');
                    }};
                    audio.onpause = function() {{
                        btn.classList.remove('playing');
                    }};
                    audio.onplay = function() {{
                        btn.classList.add('playing');
                    }};
                }}
            </script>
        </head>
        <body>
            <div class="container">
                <div class="sidebar">
                    {sidebar_header_div}
                    <div id="chat-list" style="flex:1; overflow-y:auto;">
                        {sidebar_html}
                    </div>
                    <div id="search-results" class="search-results-container">
                        <div class="search-header">
                            <span>RISULTATI RICERCA</span>
                            <span class="close-search" onclick="closeSearch()">CHIUDI</span>
                        </div>
                        <div id="search-results-content"></div>
                    </div>
                </div>
                <div class="main">
                    {chats_html}
                </div>
            </div>
            <script>
                // Media Gallery Logic
                const allMedia = {json.dumps(all_media)};
                let currentChatMedia = null;

                function openGallery(chatId) {{
                    currentChatMedia = allMedia[chatId];
                    if(!currentChatMedia) return;
                    
                    document.getElementById('media-gallery-modal').style.display = 'flex';
                    switchTab('media');
                }}

                function closeGallery() {{
                    document.getElementById('media-gallery-modal').style.display = 'none';
                }}

                function switchTab(tab) {{
                    document.querySelectorAll('.modal-tab').forEach(t => t.classList.remove('active'));
                    document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
                    
                    document.querySelector(`.modal-tab[data-tab="${{tab}}"]`).classList.add('active');
                    document.getElementById('tab-' + tab).classList.add('active');
                    
                    if(tab === 'media') renderMedia();
                    else if(tab === 'links') renderLinks();
                    else if(tab === 'docs') renderDocs();
                }}

                function renderMedia() {{
                    const grid = document.getElementById('gallery-grid');
                    grid.innerHTML = '';
                    const images = currentChatMedia.images || [];
                    const videos = currentChatMedia.videos || [];
                    
                    if(images.length === 0 && videos.length === 0) {{
                        grid.innerHTML = '<div style="grid-column:1/-1; text-align:center; padding:20px; color:#888;">Nessun media trovato.</div>';
                        return;
                    }}
                    
                    // Images
                    images.forEach(item => {{
                        const div = document.createElement('div');
                        div.className = 'gallery-item';
                        div.onclick = function() {{ openLightbox(item.src, 'image', item.name); }};
                        div.innerHTML = `<img src="${{item.src}}" loading="lazy" title="${{item.name}}">`;
                        grid.appendChild(div);
                    }});
                    
                    // Videos with play icon overlay
                    videos.forEach(item => {{
                        const div = document.createElement('div');
                        div.className = 'gallery-item video-item';
                        div.onclick = function() {{ openLightbox(item.src, 'video', item.name); }};
                        div.innerHTML = `<video src="${{item.src}}" preload="metadata"></video><div class="video-play-overlay">▶</div>`;
                        grid.appendChild(div);
                    }});
                }}

                function renderLinks() {{
                    const list = document.getElementById('link-list');
                    list.innerHTML = '';
                    const items = currentChatMedia.links || [];
                    
                    if(items.length === 0) {{
                        list.innerHTML = '<div style="text-align:center; padding:20px; color:#888;">Nessun link trovato.</div>';
                        return;
                    }}
                    
                    items.forEach(item => {{
                        const li = document.createElement('li');
                        li.innerHTML = `<a href="${{item.url}}" target="_blank">${{item.url}}</a>`;
                        list.appendChild(li);
                    }});
                }}
                
                function renderDocs() {{
                    const list = document.getElementById('docs-list');
                    list.innerHTML = '';
                    const items = currentChatMedia.docs || [];
                    
                    if(items.length === 0) {{
                        list.innerHTML = '<div style="text-align:center; padding:20px; color:#888;">Nessun documento trovato.</div>';
                        return;
                    }}
                    
                    const icons = {{'pdf': '📕', 'doc': '📘', 'docx': '📘', 'xls': '📗', 'xlsx': '📗', 'ppt': '📙', 'pptx': '📙', 'txt': '📄', 'csv': '📊', 'zip': '📦', 'rar': '📦'}};
                    
                    items.forEach(item => {{
                        const li = document.createElement('li');
                        const icon = icons[item.ext] || '📎';
                        li.innerHTML = `<span class="doc-icon">${{icon}}</span> <a href="${{item.src}}" target="_blank">${{item.name}}</a>`;
                        list.appendChild(li);
                    }});
                }}
                
                // Lightbox
                function openLightbox(src, type, name) {{
                    const lb = document.getElementById('lightbox-overlay');
                    const content = document.getElementById('lightbox-content');
                    
                    if(type === 'image') {{
                        content.innerHTML = `<img src="${{src}}" alt="${{name}}">`;
                    }} else if(type === 'video') {{
                        content.innerHTML = `<video src="${{src}}" controls autoplay style="max-width:90vw; max-height:80vh;"></video>`;
                    }}
                    
                    lb.style.display = 'flex';
                }}
                
                function closeLightbox() {{
                    const lb = document.getElementById('lightbox-overlay');
                    const content = document.getElementById('lightbox-content');
                    content.innerHTML = '';
                    lb.style.display = 'none';
                }}
                
                let currentChatId = "";
                const originalOpen = openGallery;
                openGallery = function(cid) {{ currentChatId = cid; originalOpen(cid); }};      
            </script>
            <style>
            .modal-overlay {{ position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.8); display: flex; justify-content: center; align-items: center; z-index: 1000; backdrop-filter: blur(5px); }}
            .modal-content {{ background: white; width: 80%; max-width: 900px; height: 80%; border-radius: 12px; display: flex; flex-direction: column; overflow: hidden; box-shadow: 0 10px 25px rgba(0,0,0,0.5); }}
            .modal-header {{ padding: 15px 20px; border-bottom: 1px solid #eee; display: flex; justify-content: space-between; align-items: center; background: #f9f9f9; }}
            .modal-header h3 {{ margin: 0; font-size: 18px; color: #333; }}
            .close-modal {{ font-size: 24px; cursor: pointer; color: #666; }} .close-modal:hover {{ color: #000; }}
            .modal-tabs {{ display: flex; background: #f0f0f0; border-bottom: 1px solid #ddd; }}
            .modal-tab {{ flex: 1; padding: 12px; text-align: center; cursor: pointer; border-bottom: 3px solid transparent; font-weight: 500; color: #666; font-size: 13px; }}
            .modal-tab.active {{ border-bottom-color: #2c6bed; color: #2c6bed; background: white; }}
            .modal-body {{ flex: 1; overflow-y: auto; padding: 20px; background: #fff; }}
            .tab-content {{ display: none; }} .tab-content.active {{ display: block; }}
            .gallery-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(100px, 1fr)); gap: 8px; }}
            .gallery-item {{ aspect-ratio: 1; background: #eee; border-radius: 6px; overflow: hidden; cursor: pointer; position: relative; border: 1px solid #ddd; }}
            .gallery-item img, .gallery-item video {{ width: 100%; height: 100%; object-fit: cover; transition: transform 0.2s; }}
            .gallery-item:hover img, .gallery-item:hover video {{ transform: scale(1.05); }}
            .video-item .video-play-overlay {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); background: rgba(0,0,0,0.6); color: white; width: 36px; height: 36px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 14px; pointer-events: none; }}
            .link-list, .docs-list {{ list-style: none; padding: 0; margin: 0; }}
            .link-list li, .docs-list li {{ padding: 12px 10px; border-bottom: 1px solid #eee; word-break: break-all; display: flex; align-items: center; gap: 10px; }}
            .link-list li a, .docs-list li a {{ color: #2c6bed; text-decoration: none; }} .link-list li a:hover, .docs-list li a:hover {{ text-decoration: underline; }}
            .doc-icon {{ font-size: 20px; }}
            
            /* Lightbox */
            #lightbox-overlay {{ position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.95); display: none; justify-content: center; align-items: center; z-index: 2000; cursor: pointer; }}
            #lightbox-content {{ max-width: 95vw; max-height: 95vh; }}
            #lightbox-content img {{ max-width: 90vw; max-height: 85vh; border-radius: 4px; box-shadow: 0 5px 30px rgba(0,0,0,0.5); }}
            #lightbox-close {{ position: fixed; top: 20px; right: 30px; font-size: 40px; color: white; cursor: pointer; z-index: 2001; }}
            </style>
            
            <div id="media-gallery-modal" class="modal-overlay" style="display:none;">
                <div class="modal-content">
                    <div class="modal-header">
                        <h3>📂 Media Gallery</h3>
                        <span class="close-modal" onclick="closeGallery()">&times;</span>
                    </div>
                    <div class="modal-tabs">
                        <div class="modal-tab active" data-tab="media" onclick="switchTab('media')">📷 Media</div>
                        <div class="modal-tab" data-tab="links" onclick="switchTab('links')">🔗 Link</div>
                        <div class="modal-tab" data-tab="docs" onclick="switchTab('docs')">📎 Documenti</div>
                    </div>
                    <div class="modal-body">
                        <div id="tab-media" class="tab-content active">
                            <div id="gallery-grid" class="gallery-grid"></div>
                        </div>
                        <div id="tab-links" class="tab-content">
                            <ul id="link-list" class="link-list"></ul>
                        </div>
                        <div id="tab-docs" class="tab-content">
                            <ul id="docs-list" class="docs-list"></ul>
                        </div>
                    </div>
                </div>
            </div>
            
            <div id="lightbox-overlay" onclick="closeLightbox()">
                <span id="lightbox-close">&times;</span>
                <div id="lightbox-content"></div>
            </div>
            
        </body>
        </html>
        """
        
        path = os.path.join(self.output_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            f.write(full_html)
        print(f"Generated: {path}")

# ==========================================
# MAIN LOGIC
# ==========================================

def build_chat_attachment_map(filepath):
    """
    Reads 'Chat' sheet to build a map of Timestamp -> Attachment Filename.
    Used to resolve attachments in 'Instant Messages'.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if 'Chat' not in wb.sheetnames: return {}
        sheet = wb['Chat']
        rows = list(sheet.iter_rows(values_only=True))
        
        # Simple/Safe header detection
        header = None
        start_row = 0
        for i, row in enumerate(rows[:5]):
            r_str = [str(x).lower() for x in row if x]
            # Key checks: 'timestamp: ora' (split), 'corpo', 'allegato'
            matches = 0
            if any("timestamp" in x for x in r_str): matches += 1
            if any("corpo" in x for x in r_str) or any("body" in x for x in r_str): matches += 1
            if matches >= 2:
                header = row
                start_row = i + 1
                break
        
        if not header: return {}
        
        col_map = {str(n).lower(): idx for idx, n in enumerate(header) if n}
        
        # Identify indices
        c_time_idx = -1
        # 'Timestamp: Ora' usually contains the full date/time string "20/05/2025 21:45:46(UTC+1)"
        for k, v in col_map.items():
            if "timestamp" in k and "ora" in k: c_time_idx = v; break
        
        c_att_idx = -1
        for k, v in col_map.items():
            if "allegato" in k and "#1" in k and "dettagli" not in k: c_att_idx = v; break
        
        if c_time_idx == -1 or c_att_idx == -1: return {}
        
        lookup = {}
        for row in rows[start_row:]:
            if len(row) <= max(c_time_idx, c_att_idx): continue
            
            ts_val = str(row[c_time_idx]) if row[c_time_idx] else ""
            att_val = str(row[c_att_idx]) if row[c_att_idx] else ""
            
            if ts_val and att_val:
                # Clean timestamp: remove newlines, (UTC...), whitespace
                ts_clean = re.sub(r'\(UTC.*?\)', '', ts_val).strip()
                lookup[ts_clean] = att_val
                
        return lookup
    except Exception as e:
        print(f"Error building attachment map: {e}")
        return {}

def run_generation(input_file, style="signal", source_type="chats", output_dir=None, format_type="auto"):
    input_path = os.path.abspath(input_file)
    base_dir = os.path.dirname(input_path)
    
    if output_dir:
        out_dir = output_dir
    else:
        out_dir = os.path.join(base_dir, "VisualReport")
        
    os.makedirs(out_dir, exist_ok=True)
    
    print(f"Input: {input_path}")
    print(f"Output: {out_dir}")
    print(f"Style: {style}")
    print(f"Source Mode: {source_type}")

    # 1. Determine Parser based on Source Type
    parser_impl = None
    att_lookup = None
    
    if source_type == "instant":
        print("Using Instant Messages Parser...")
        # Build attachment lookup from 'Chat' tab
        print("Building Attachment Lookup from 'Chat' tab...")
        att_lookup = build_chat_attachment_map(input_path)
        print(f"Attachment Lookup Built: {len(att_lookup)} entries.")
        
        parser_impl = WhatsAppParser() # Uses 'Instant Messages' tab logic
        chats = parser_impl.parse(input_path, attachment_lookup=att_lookup)
    else:
        print("Using Standard Chat Parser...")
        parser_impl = CellebriteParser() # Uses 'Chat' tab logic
        chats = parser_impl.parse(input_path)
    print(f"Parsed {len(chats)} chats.")
    
    # 3. Scan attachments
    file_map = process_attachments(base_dir, out_dir)
    
    # 4. Render
    css = CSS_SIGNAL if style == "signal" else CSS_WHATSAPP
    renderer = HTMLRenderer(css, file_map, out_dir, style_mode=style)
    renderer.render(chats)
    print("Generation Complete.")
    return out_dir

# ==========================================
# GUI CLASSES
# ==========================================

class RedirectText:
    def __init__(self, text_ctrl):
        self.output = text_ctrl

    def write(self, string):
        self.output.insert(tk.END, string)
        self.output.see(tk.END)

    def flush(self):
        pass

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Chat Report Generator (Lite)")
        self.root.geometry("600x500")
        
        # Styles
        self.root.configure(bg="#f0f0f0")
        font_label = ("Segoe UI", 10)
        
        # 1. File Selection
        frame_file = tk.Frame(root, bg="#f0f0f0", pady=10)
        frame_file.pack(fill=tk.X, padx=20)
        
        tk.Label(frame_file, text="File Excel:", bg="#f0f0f0", font=font_label).pack(anchor="w")
        
        self.entry_file = tk.Entry(frame_file, width=50)
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=5)
        
        btn_browse = tk.Button(frame_file, text="Sfoglia...", command=self.browse_file, font=font_label)
        btn_browse.pack(side=tk.RIGHT, padx=5)
        
        # 2. Options
        frame_opts = tk.Frame(root, bg="#f0f0f0", pady=10)
        frame_opts.pack(fill=tk.X, padx=20)
        
        # Style
        tk.Label(frame_opts, text="Stile Grafico:", bg="#f0f0f0", font=font_label).grid(row=0, column=0, padx=5, sticky="w")
        self.var_style = tk.StringVar(value="signal")
        tk.Radiobutton(frame_opts, text="Signal (Blue)", variable=self.var_style, value="signal", bg="#f0f0f0").grid(row=0, column=1)
        tk.Radiobutton(frame_opts, text="WhatsApp (Green)", variable=self.var_style, value="whatsapp", bg="#f0f0f0").grid(row=0, column=2)
        
        # Source
        tk.Label(frame_opts, text="Sorgente:", bg="#f0f0f0", font=font_label).grid(row=1, column=0, padx=5, sticky="w", pady=5)
        self.var_source = tk.StringVar(value="chats")
        tk.Radiobutton(frame_opts, text="Chat (Standard)", variable=self.var_source, value="chats", bg="#f0f0f0").grid(row=1, column=1, sticky="w")
        tk.Radiobutton(frame_opts, text="Messaggi Istantanei", variable=self.var_source, value="instant", bg="#f0f0f0").grid(row=1, column=2, sticky="w")
        
        # 3. Action
        frame_action = tk.Frame(root, bg="#f0f0f0", pady=20)
        frame_action.pack(fill=tk.X, padx=20)
        
        btn_run = tk.Button(frame_action, text="GENERA REPORT", command=self.start_generation, 
                            bg="#2c6bed", fg="white", font=("Segoe UI", 12, "bold"), height=2)
        btn_run.pack(fill=tk.X, pady=(0, 10))
        
        self.btn_open = tk.Button(frame_action, text="APRI CARTELLA OUTPUT", command=self.open_output, 
                                  state=tk.DISABLED, font=("Segoe UI", 10))
        self.btn_open.pack(fill=tk.X)
        
        # 4. Log Output
        tk.Label(root, text="Log:", bg="#f0f0f0", font=font_label).pack(anchor="w", padx=20)
        self.text_log = scrolledtext.ScrolledText(root, height=15)
        self.text_log.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Redirect stdout
        sys.stdout = RedirectText(self.text_log)
        sys.stderr = RedirectText(self.text_log)
        
        self.last_output_dir = None

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if filename:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, filename)

    def start_generation(self):
        input_path = self.entry_file.get()
        if not input_path or not os.path.exists(input_path):
            messagebox.showwarning("Attenzione", "Seleziona un file Excel valido.")
            return
            
        style = self.var_style.get()
        source = self.var_source.get()
        
        # Disable button
        self.btn_open.config(state=tk.DISABLED)
        
        t = threading.Thread(target=self.run_process, args=(input_path, style, source))
        t.start()
        
    def run_process(self, input_path, style, source):
        print("--- Inizio Elaborazione ---")
        try:
            # Determine output automatically (same dir)
            out_dir = run_generation(input_path, style=style, source_type=source, output_dir=None)
            self.last_output_dir = out_dir
            
            self.root.after(0, lambda: self.finish_success())
        except Exception as e:
            error_msg = str(e)
            print(f"ERRORE: {error_msg}")
            self.root.after(0, lambda: messagebox.showerror("Errore", f"Si è verificato un errore:\n{error_msg}"))
        finally:
            print("--- Fine Elaborazione ---")

    def finish_success(self):
        messagebox.showinfo("Successo", "Report generato con successo!")
        self.btn_open.config(state=tk.NORMAL)
        
    def open_output(self):
        if self.last_output_dir and os.path.exists(self.last_output_dir):
            os.startfile(self.last_output_dir)
        else:
            messagebox.showwarning("Errore", "Cartella non trovata.")

# ==========================================
# ENTRY POINT
# ==========================================

def main():
    # If args passed, run CLI mode
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="Convert Excel Chat export to HTML")
        parser.add_argument("input_file", help="Path to Excel file")
        parser.add_argument("--style", choices=["signal", "whatsapp"], default="signal", help="Output visual style")
        parser.add_argument("--output", help="Output directory")
        parser.add_argument("--format", choices=["auto", "cellebrite", "whatsapp"], default="auto", help="Input format")
        
        args = parser.parse_args()
        run_generation(args.input_file, args.style, args.output, args.format)
    else:
        # GUI Mode
        root = tk.Tk()
        app = App(root)
        root.mainloop()

if __name__ == "__main__":
    main()
